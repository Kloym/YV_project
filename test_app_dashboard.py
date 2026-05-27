import patch_win
import os
import sqlite3
import webbrowser
import io
import logging
from threading import Timer
import dash
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import dash_bootstrap_components as dbc
from dash import html, dcc, Input, Output, State
import duckdb
import dash_ag_grid as dag
from openpyxl.styles import PatternFill
from flask_caching import Cache
import cProfile
import pstats
import io
from functools import wraps



CUSTOM_ICON = "data:image/svg+xml;base64,PD94bWwgdmVyc2lvbj0iMS4wIiBlbmNvZGluZz0iVVRGLTgiPz4gPHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHhtbG5zOnhsaW5rPSJodHRwOi8vd3d3LnczLm9yZy8xOTk5L3hsaW5rIiB2ZXJzaW9uPSIxLjEiIGlkPSLQodC70L7QuV8xIiB4PSIwcHgiIHk9IjBweCIgdmlld0JveD0iMCAwIDE5Ni4xNyA0Mi4xNiIgc3R5bGU9ImVuYWJsZS1iYWNrZ3JvdW5kOm5ldyAwIDAgMTk2LjE3IDQyLjE2OyIgeG1sOnNwYWNlPSJwcmVzZXJ2ZSI+IDxzdHlsZSB0eXBlPSJ0ZXh0L2NzcyI+IC5zdDB7ZmlsbDojQzczMDQ3O30gPC9zdHlsZT4gPGc+IDxwYXRoIGNsYXNzPSJzdDAiIGQ9Ik03Mi4yOSwyNC40YzAsMCwzLjA4LDAsNi00LjEzYzEuODYtMi42OCw0LjMtMy43Myw2LjU3LTIuNTljLTAuNTcsMS4xMy0yLjE5LDMuMDgtMy45Nyw0LjIyIGMxLjU0LTAuNTcsMy40OS0yLjI3LDQuNjItMy43M2MwLjY1LDAuODEsMC42NSwyLjExLTAuNTcsMy42NUM4My40NywyMy41MSw3OC4yLDI2LjAyLDcyLjI5LDI0LjQgTTkxLjE4LDYuMjRWMi40MyBjMC0xLjMtMS4wNS0yLjQzLTIuNDMtMi40M0g3My4wMmMtMS4zLDAtMi40MywxLjA1LTIuNDMsMi40M3YzLjQxdjIuNTljMCwxLjMtMS4wNSwyLjQzLTIuNDMsMi40M2gtMi4wM2gtMy44OSBjLTEuMywwLTIuNDMsMS4wNS0yLjQzLDIuNDN2My41N2M0LjQ2LDMsOS44MS0xLjg2LDEzLjc4LTMuNjVjNS41OS0yLjQzLDEzLjg2LTAuMDgsMTQuNTEsMC4xNmMwLjY1LDAuMjQsMS45NSwxLjMtNC45NSwxLjcgYy02Ljg5LDAuMzItMTAuODYsNC4zLTEwLjA1LDMuNDljMC44MS0wLjczLDkuMzItMy44MSw0LjQ2LTAuNjVjLTQuNzgsMy4xNi03LjEzLDUuNTktMTAuMDUsNS43NmMtMi45MiwwLjE2LTQuMzgtMC44MS02LjA4LTEuMDUgYy0xLjIyLTAuMTYtMS41NCwwLjU3LTEuNjIsMS4wNXY1LjM1YzAsMS4zLDEuMDUsMi40MywyLjQzLDIuNDNoMy41N2gyLjM1YzEuMywwLDIuNDMsMS4wNSwyLjQzLDIuNDN2MS43OHY0LjA1IGMwLDEuMywxLjA1LDIuNDMsMi40MywyLjQzaDE1LjczYzEuMywwLDIuNDMtMS4wNSwyLjQzLTIuNDN2LTMuNTd2LTIuMzVjMC0xLjMsMS4wNS0yLjQzLDIuNDMtMi40M2gyLjY4aDMuMjQgYzEuMywwLDIuNDMtMS4wNSwyLjQzLTIuNDN2LTMuNjVjLTQuNDYtMy05LjgxLDEuODYtMTMuNzgsMy42NWMtNS41OSwyLjQzLTEzLjg2LDAuMDgtMTQuNTEtMC4xNmMtMC42NS0wLjI0LTEuOTUtMS4zOCw0Ljk1LTEuNyBjNi44OS0wLjMyLDEwLjg2LTQuMywxMC4wNS0zLjQ5Yy0wLjgxLDAuNzMtOS4zMiwzLjgxLTQuNDYsMC42NWM0Ljg2LTMuMTYsNy4xMy01LjUxLDEwLjEzLTUuNzZjMi45Mi0wLjE2LDQuMzgsMC44MSw2LjA4LDEuMDUgYzEuMjIsMC4xNiwxLjU0LTAuNTcsMS42Mi0xLjA1di01LjI3YzAtMS4zLTEuMDUtMi40My0yLjQzLTIuNDNoLTMuOTdoLTEuOTVjLTEuMywwLTIuNDMtMS4wNS0yLjQzLTIuNDNMOTEuMTgsNi4yNHoiPjwvcGF0aD4gPGc+IDxwYXRoIGNsYXNzPSJzdDAiIGQ9Ik00My4zMiwyLjc1Yy0wLjg2LTAuNjItMS43My0xLjEzLTIuNjItMS41NGMwLDAuOTgsMCwyLjE0LDAsMy4zNWMwLjI5LDAuMTgsMC41OCwwLjM3LDAuODcsMC41NyBjMy4xOCwyLjI4LDYuNDUsOC4yOSwyLjQyLDE1LjQ2Yy0xLjI4LDIuMjgtMi4zMSwzLjYzLTMuNDYsNC45NWMtMS41NCwxLjc3LTIuOTQsMy4yNS01LjksNS45NWMtMy4wNywyLjgtNy42MSw1LjY3LTEwLjIsNy4yMiBjLTIuNTktMS41NS03LjE0LTQuNDItMTAuMi03LjIyYy0yLjk2LTIuNy00LjM1LTQuMTgtNS45LTUuOTVjLTEuMTUtMS4zMi0yLjE4LTIuNjctMy40Ni00Ljk1QzAuODMsMTMuNDMsNC4xLDcuNDMsNy4yOCw1LjE0IGMwLjI5LTAuMjEsMC41OC0wLjQsMC44Ny0wLjU3YzAtMS4yMSwwLTIuMzcsMC0zLjM1QzcuMjcsMS42Miw2LjQsMi4xMyw1LjUzLDIuNzVjLTQuMDQsMi45LTguMjUsMTAuNDMtMy4yNywxOS4yOSBjMS4yLDIuMTQsMi4zLDMuNzEsMy44LDUuNDRjMS42MiwxLjg2LDMuMDcsMy40LDYuMTMsNi4yYzQuMDUsMy43LDEwLjI5LDcuMzgsMTIuMjIsOC40OGgwYzEuOTMtMS4xLDguMTctNC43OCwxMi4yMi04LjQ4IGMzLjA2LTIuOCw0LjUyLTQuMzQsNi4xMy02LjJjMS41LTEuNzMsMi42LTMuMywzLjgtNS40NEM1MS41NiwxMy4xOSw0Ny4zNiw1LjY2LDQzLjMyLDIuNzUiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTQzLjA1LDEwLjQ1Yy0wLjU2LTEuMzYtMS40My0yLjQzLTIuMzUtMy4yNWMwLDIuNjcsMCw1LjExLDAsNS45MmMwLDAuMTgtMC4wMSwwLjM1LTAuMDIsMC41MyBjMC4wMSwwLjc4LTAuMTIsMS42Mi0wLjQsMi41Yy0wLjI4LDEtMC43LDEuODktMS4xOCwyLjcxYzAsMCwwLDAsMCwwLjAxYy0wLjEzLDAuMjQtMC4zMiwwLjUyLTAuNTMsMC44MyBjLTAuMDYsMC4wOS0wLjEzLDAuMTgtMC4xOSwwLjI3Yy0wLjAzLDAuMDQtMC4wNSwwLjA3LTAuMDgsMC4xMWMtMC4xLDAuMTQtMC4yMSwwLjI4LTAuMzIsMC40MmMwLDAuMDEtMC4wMSwwLjAxLTAuMDEsMC4wMiBjLTAuMDksMC4xMS0wLjE4LDAuMjMtMC4yNiwwLjM0Yy0wLjg5LDEuMTItMi4wNSwyLjQzLTMuMywzLjcyYy0wLjYsMC42NC0xLjIzLDEuMjYtMS44NiwxLjg1Yy0wLjIzLDAuMjItMC40NiwwLjQzLTAuNjksMC42NCBjLTIuNzQsMi41MS01LjYxLDQuNDgtNy40Myw1LjYzYy0xLjgyLTEuMTUtNC42OS0zLjEyLTcuNDMtNS42M2MtMC4yMy0wLjIxLTAuNDYtMC40Mi0wLjY5LTAuNjRjLTAuNjQtMC42LTEuMjYtMS4yMS0xLjg2LTEuODUgYy0xLjI1LTEuMjktMi40MS0yLjYxLTMuMy0zLjcyYy0wLjA5LTAuMTEtMC4xOC0wLjIyLTAuMjYtMC4zNGMwLTAuMDEtMC4wMS0wLjAxLTAuMDEtMC4wMmMtMC4xMS0wLjE0LTAuMjEtMC4yOC0wLjMyLTAuNDIgYy0wLjAzLTAuMDQtMC4wNS0wLjA3LTAuMDgtMC4xMWMtMC4wNi0wLjA5LTAuMTMtMC4xOC0wLjE5LTAuMjdjLTAuMjEtMC4zMS0wLjQtMC41OS0wLjUzLTAuODNjMCwwLDAsMCwwLTAuMDEgYy0wLjQ4LTAuODItMC45LTEuNzEtMS4xOC0yLjcxYy0wLjI3LTAuODktMC40MS0xLjcyLTAuNC0yLjVjLTAuMDEtMC4xNy0wLjAyLTAuMzUtMC4wMi0wLjUzYzAtMC44MSwwLTMuMjYsMC01LjkyIEM3LjIzLDguMDIsNi4zNiw5LjA5LDUuOCwxMC40NWMtMS4xNywyLjg0LTAuNzIsNi4xNSwxLjMzLDkuODNjMS4zMSwyLjM0LDUuMDEsNi4zNiw3LjgzLDguOTRjMy45NywzLjY1LDguMTIsNi4xNiw5LjQ3LDYuOTRoMCBjMS4zNS0wLjc4LDUuNDktMy4zLDkuNDctNi45NGMyLjgyLTIuNTksNi41Mi02LjYsNy44My04Ljk0QzQzLjc4LDE2LjU5LDQ0LjIyLDEzLjI5LDQzLjA1LDEwLjQ1Ij48L3BhdGg+IDxwYXRoIGNsYXNzPSJzdDAiIGQ9Ik0zOC4zMSwwLjM3Yy0yLjM2LTAuNTctNC44OS0wLjQ4LTcuNzcsMC4yNmMtMi41NywwLjY2LTQuNiwxLjkzLTYuMTEsMy4yMWMtMS41MS0xLjI4LTMuNTQtMi41NS02LjExLTMuMjEgYy0yLjg4LTAuNzQtNS40MS0wLjgzLTcuNzctMC4yNmMtMC4wNSwwLjAxLTAuMTcsMC4wNC0wLjE3LDAuMDRzMCwzLjUyLDAsMTIuNjljMCwzLDEuMzgsNS4wNSwyLjk1LDYuODlWMi45NyBjMS4zLTAuMDUsMi43LDAuMTIsNC4yNCwwLjUyYzMuNDcsMC44OSw1Ljc4LDMuMjQsNi44Niw0LjU3bDAuMDEsMC4wMWwwLjAxLTAuMDFjMS4wOC0xLjMzLDMuMzgtMy42OCw2Ljg1LTQuNTcgYzEuNTQtMC40LDIuOTMtMC41Niw0LjI0LTAuNTJWMjBjMS41Ny0xLjg0LDIuOTUtMy44OSwyLjk1LTYuODljMC05LjE3LDAtMTIuNjksMC0xMi42OVMzOC4zNiwwLjM4LDM4LjMxLDAuMzciPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTMxLDUuOTNjLTIuMjgsMC45OC01LjE0LDQuMDEtNi41Nyw1Ljk4QzIzLDkuOTQsMjAuMTQsNi45MSwxNy44NSw1LjkzYy0wLjY5LTAuMy0xLjQ5LTAuNDctMi4yOS0wLjU4IGwwLjAxLDE3LjExYzAuOTMsMS4wMiwxLjkxLDIsMi45NCwyLjkyVjkuOTRjMi4zMSwxLjUyLDQuMTksNC4zLDUuODcsNi42NGwwLjA1LDAuMDhsMC4wNS0wLjA4YzEuNjgtMi4zNCwzLjU0LTUuMTIsNS44Ni02LjY0IHYxNS40NWMxLjAyLTAuOTIsMi4wMS0xLjksMi45NC0yLjkybDAuMDEtMTcuMTFDMzIuNDksNS40NiwzMS42OSw1LjYzLDMxLDUuOTMiPjwvcGF0aD4gPC9nPiA8Zz4gPGc+IDxwYXRoIGNsYXNzPSJzdDAiIGQ9Ik0xMTIuOTksMC4xM2g0Ljc2djEuNzZoLTIuODl2MS40OWMyLjYxLDAsMy41NCwxLjQ5LDMuNTQsMy4xMWMwLDEuOTktMS4zNSwzLjExLTMuNTQsMy4xMWgtMS44NlYwLjEzeiBNMTE2LjQ4LDYuNDljMC0wLjg4LTAuNTctMS4zOC0xLjYyLTEuMzh2Mi43NkMxMTUuOTEsNy44NiwxMTYuNDgsNy4zNiwxMTYuNDgsNi40OXoiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTExOS4zNCw0Ljg2YzAtMi45MSwyLjAxLTQuODYsNC43OC00Ljg2YzIuNzcsMCw0LjgxLDEuOTYsNC44MSw0Ljg2YzAsMi43Ny0yLjA0LDQuODYtNC44MSw0Ljg2IEMxMjEuMzUsOS43MywxMTkuMzQsNy42NCwxMTkuMzQsNC44NnogTTEyNi45Nyw0Ljg2YzAtMS43Mi0xLjE1LTIuOTctMi44NC0yLjk3cy0yLjg0LDEuMjYtMi44NCwyLjk3czEuMTUsMi45NywyLjg0LDIuOTcgUzEyNi45Nyw2LjU4LDEyNi45Nyw0Ljg2eiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTMwLjQyLDYuOTVsMC4yMy02LjgxaDQuODZ2OS40NmgtMS44MlYxLjg2aC0xLjI4bC0wLjE2LDUuMDhjLTAuMDcsMi4zMS0xLjE1LDIuNzItMi43NywyLjcyVjcuOTEgQzEzMC4wMSw3LjkxLDEzMC4zOSw3LjYxLDEzMC40Miw2Ljk1eiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTM3LjEzLDAuMTNoMS44OXYzLjI0YzIuNjUsMCwzLjU4LDEuNDksMy41OCwzLjExYzAsMS45OS0xLjM1LDMuMTEtMy41OCwzLjExaC0xLjg5VjAuMTN6IE0xNDAuNjUsNi40OSBjMC0wLjg4LTAuNTctMS4zNS0xLjYyLTEuMzV2Mi43QzE0MC4wOCw3Ljg0LDE0MC42NSw3LjM2LDE0MC42NSw2LjQ5eiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTQ3LjU0LDUuNTFoLTEuODl2NC4wOGgtMS44OVYwLjEzaDEuODl2My42MmgxLjg5VjAuMTNoMS44OXY5LjQ2aC0xLjg5VjUuNTF6Ij48L3BhdGg+IDxwYXRoIGNsYXNzPSJzdDAiIGQ9Ik0xNTEuMDUsMC4xM2gxLjc2VjQuOGwzLjUzLTQuOGgwLjI3djkuNTloLTEuNzZWNC45M2wtMy41Myw0LjhoLTAuMjdWMC4xM3oiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTE1OC4yMiwwLjEzaDEuODl2Ny43aDEuODl2LTcuN2gxLjg5djcuN2gwLjU0djMuMjRoLTEuNzZWOS41OWgtNC40NlYwLjEzeiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTY5LjI1LDBoMC4yN2w0LjM0LDkuNTloLTIuMDFsLTAuNTktMS40MmgtMy43OGwtMC41OSwxLjQyaC0xLjk2TDE2OS4yNSwweiBNMTcwLjUzLDYuNDlsLTEuMTgtMi44MSBsLTEuMTgsMi44MUgxNzAuNTN6Ij48L3BhdGg+IDwvZz4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTExMi45OSwxNi4zNWgxLjc2djQuNjZsMy41My00LjhoMC4yN3Y5LjU5aC0xLjc2di00LjY2bC0zLjUzLDQuOGgtMC4yN1YxNi4zNXoiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTEyNC4yNSwxOC45NGMxLjEyLTEuMzYsMi4zMS0yLjM4LDQuMDctMi41OXY5LjQ2aC0xLjgydi02LjcybC0yLjE5LDIuMzFoLTAuMTRsLTIuMTgtMi4zMnY2LjczaC0xLjgydi05LjQ2IEMxMjEuOTIsMTYuNTYsMTIzLjExLDE3LjU4LDEyNC4yNSwxOC45NHoiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTEyOS45MywxNi4zNWg0LjM5djEuNzZoLTIuNXYxLjg4aDIuMDN2MS43NmgtMi4wM3YyLjMxaDIuN3YxLjc2aC00LjU5VjE2LjM1eiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTM5LjUyLDIxLjczaC0xLjg5djQuMDhoLTEuODl2LTkuNDZoMS44OXYzLjYyaDEuODl2LTMuNjJoMS44OXY5LjQ2aC0xLjg5VjIxLjczeiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTQzLjA0LDE2LjM1aDEuNzZ2NC42NmwzLjUzLTQuOGgwLjI3djkuNTloLTEuNzZ2LTQuNjZsLTMuNTMsNC44aC0wLjI3VjE2LjM1eiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTEyLjk5LDMyLjU2aDEuNDljMi4zLDAsMy40NywwLjkyLDMuNDcsMi41NGMwLDAuODItMC4zNCwxLjQ5LTEuMDUsMS44NWMxLjA1LDAuMzksMS40NiwxLjIsMS40NiwyLjMxIGMwLDEuNzYtMS4yMiwyLjc2LTMuNjEsMi43NmgtMS43NlYzMi41NnogTTExNi4xNiwzNS4yNGMwLTAuNzQtMC41LTEuMDQtMS40Mi0xLjA0djIuMUMxMTUuNjUsMzYuMjksMTE2LjE2LDM1LjkxLDExNi4xNiwzNS4yNHogTTExNi41NywzOS4xM2MwLTAuODUtMC41NC0xLjI0LTEuODItMS4yNHYyLjQ5QzExNi4wMiw0MC4zNywxMTYuNTcsMzkuOTgsMTE2LjU3LDM5LjEzeiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTE5LjEsNDAuOTdjMC0wLjY2LDAuNDctMS4xOCwxLjE2LTEuMThjMC42OCwwLDEuMTUsMC41MSwxLjE1LDEuMThjMCwwLjY2LTAuNDcsMS4xOS0xLjE1LDEuMTkgQzExOS41Nyw0Mi4xNiwxMTkuMSw0MS42MywxMTkuMSw0MC45N3oiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTEyMi42MiwzMi41NmgxLjQ5YzIuMywwLDMuNDcsMC45MiwzLjQ3LDIuNTRjMCwwLjgyLTAuMzQsMS40OS0xLjA1LDEuODVjMS4wNSwwLjM5LDEuNDYsMS4yLDEuNDYsMi4zMSBjMCwxLjc2LTEuMjIsMi43Ni0zLjYxLDIuNzZoLTEuNzZWMzIuNTZ6IE0xMjUuOCwzNS4yNGMwLTAuNzQtMC41LTEuMDQtMS40Mi0xLjA0djIuMUMxMjUuMjgsMzYuMjksMTI1LjgsMzUuOTEsMTI1LjgsMzUuMjR6IE0xMjYuMiwzOS4xM2MwLTAuODUtMC41NC0xLjI0LTEuODItMS4yNHYyLjQ5QzEyNS42NSw0MC4zNywxMjYuMiwzOS45OCwxMjYuMiwzOS4xM3oiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTEyOC43Myw0MC45N2MwLTAuNjYsMC40Ny0xLjE4LDEuMTYtMS4xOGMwLjY4LDAsMS4xNSwwLjUxLDEuMTUsMS4xOGMwLDAuNjYtMC40NywxLjE5LTEuMTUsMS4xOSBDMTI5LjIsNDIuMTYsMTI4LjczLDQxLjYzLDEyOC43Myw0MC45N3oiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTEzNC45NiwzMi41NmgxLjQ5YzIuMywwLDMuNDcsMC45MiwzLjQ3LDIuNTRjMCwwLjgyLTAuMzQsMS40OS0xLjA1LDEuODVjMS4wNSwwLjM5LDEuNDYsMS4yLDEuNDYsMi4zMSBjMCwxLjc2LTEuMjIsMi43Ni0zLjYxLDIuNzZoLTEuNzZWMzIuNTZ6IE0xMzguMTMsMzUuMjRjMC0wLjc0LTAuNS0xLjA0LTEuNDItMS4wNHYyLjFDMTM3LjYyLDM2LjI5LDEzOC4xMywzNS45MSwxMzguMTMsMzUuMjR6IE0xMzguNTQsMzkuMTNjMC0wLjg1LTAuNTQtMS4yNC0xLjgyLTEuMjR2Mi40OUMxMzcuOTgsNDAuMzcsMTM4LjU0LDM5Ljk4LDEzOC41NCwzOS4xM3oiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTE0MS40NywzMi41Nmg0LjM5djEuNzZoLTIuNXYxLjg4aDIuMDN2MS43NmgtMi4wM3YyLjMxaDIuN3YxLjc2aC00LjU5VjMyLjU2eiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTQ3LjI4LDMyLjU2aDEuNDljMi41NywwLDMuOTksMS4wNSwzLjk5LDMuMTFjMCwxLjYyLTAuOTMsMy4xMS0zLjU4LDMuMTF2My4yNGgtMS44OVYzMi41NnogTTE1MC43OSwzNS42NyBjMC0wLjg4LTAuNTctMS4zNS0xLjYyLTEuMzV2Mi43QzE1MC4yMiwzNy4wMiwxNTAuNzksMzYuNTUsMTUwLjc5LDM1LjY3eiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTUzLjksMzIuNTZoNC4zOXYxLjc2aC0yLjV2MS44OGgyLjAzdjEuNzZoLTIuMDN2Mi4zMWgyLjd2MS43NmgtNC41OVYzMi41NnoiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTE1OS4zLDM3LjI5YzAtMi45NywyLjA4LTQuODYsNC45Mi00Ljg2YzAuMzksMCwwLjY5LDAuMDMsMSwwLjEydjEuODhjLTAuMzEtMC4wOC0wLjU4LTAuMTEtMC45OS0wLjExIGMtMS43NiwwLTIuOTcsMS4xNS0yLjk3LDIuOTdjMCwxLjgzLDEuMjIsMi45NywyLjk3LDIuOTdjMC40MSwwLDAuNjgtMC4wMywwLjk5LTAuMTF2MS44NWMtMC4zMSwwLjEyLTAuNjIsMC4xNS0xLDAuMTUgQzE2MS4zOCw0Mi4xNiwxNTkuMyw0MC4xMywxNTkuMywzNy4yOXoiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTE3MC4zLDMyLjQzaDAuMjdsNC4zNCw5LjU5aC0yLjAxbC0wLjU5LTEuNDJoLTMuNzhsLTAuNTksMS40MmgtMS45NkwxNzAuMywzMi40M3ogTTE3MS41OCwzOC45MWwtMS4xOC0yLjgxIGwtMS4xOCwyLjgxSDE3MS41OHoiPjwvcGF0aD4gPHBhdGggY2xhc3M9InN0MCIgZD0iTTE3NS43OSwzMi41Nmg0LjM5djEuNzZoLTIuNXYxLjg4aDIuMDN2MS43NmgtMi4wM3YyLjMxaDIuN3YxLjc2aC00LjU5VjMyLjU2eiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTgxLjU5LDMyLjU2aDEuNDljMi4zLDAsMy40NywwLjkyLDMuNDcsMi41NGMwLDAuODItMC4zNCwxLjQ5LTEuMDUsMS44NWMxLjA1LDAuMzksMS40NiwxLjIsMS40NiwyLjMxIGMwLDEuNzYtMS4yMiwyLjc2LTMuNjEsMi43NmgtMS43NlYzMi41NnogTTE4NC43NywzNS4yNGMwLTAuNzQtMC41LTEuMDQtMS40Mi0xLjA0djIuMUMxODQuMjYsMzYuMjksMTg0Ljc3LDM1LjkxLDE4NC43NywzNS4yNHogTTE4NS4xNywzOS4xM2MwLTAuODUtMC41NC0xLjI0LTEuODItMS4yNHYyLjQ5QzE4NC42Miw0MC4zNywxODUuMTcsMzkuOTgsMTg1LjE3LDM5LjEzeiI+PC9wYXRoPiA8cGF0aCBjbGFzcz0ic3QwIiBkPSJNMTkxLjU2LDMyLjQzaDAuMjdsNC4zNCw5LjU5aC0yLjAxbC0wLjU5LTEuNDJoLTMuNzhsLTAuNTksMS40MmgtMS45NkwxOTEuNTYsMzIuNDN6IE0xOTIuODUsMzguOTEgbC0xLjE4LTIuODFsLTEuMTgsMi44MUgxOTIuODV6Ij48L3BhdGg+IDwvZz4gPC9nPiA8L3N2Zz4g"

def profile_to_txt(filename):
    """Декоратор для записи результатов профилирования в txt-файл"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            profiler = cProfile.Profile()
            profiler.enable()
            
            result = func(*args, **kwargs)
            
            profiler.disable()
            s = io.StringIO()
            ps = pstats.Stats(profiler, stream=s).sort_stats('cumulative')
            ps.print_stats(50)
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(s.getvalue())
                
            return result
        return wrapper
    return decorator

# --- КОНФИГУРАЦИЯ И КОНСТАНТЫ ---
DB_NAME = "database.db"

MONTHS_RU = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

QUARTERS_RU = {1: "1 Квартал", 2: "2 Квартал", 3: "3 Квартал", 4: "4 Квартал"}

AG_GRID_LOCALE_RU = {
    "selectAll": "(Выбрать все)",
    "searchOoo": "Поиск...",
    "blanks": "(Пустые)",
    "noMatches": "Нет совпадений",
    "filterOoo": "Фильтр...",
    "equals": "Равно",
    "notEqual": "Не равно",
    "blank": "Пусто",
    "notBlank": "Не пусто",
    "empty": "Выберите тип фильтра",
    "lessThan": "Меньше чем",
    "greaterThan": "Больше чем",
    "lessThanOrEqual": "Меньше или равно",
    "greaterThanOrEqual": "Больше или равно",
    "inRange": "В промежутке",
    "inRangeStart": "от",
    "inRangeEnd": "до",
    "contains": "Содержит",
    "notContains": "Не содержит",
    "startsWith": "Начинается с",
    "endsWith": "Заканчивается на",
    "andCondition": "И",
    "orCondition": "ИЛИ",
    "columns": "Колонки",
    "filters": "Фильтры",
    "pivotMode": "Режим сводной таблицы",
    "groups": "Группировка строк",
    "rowGroupColumnsEmptyMessage": "Перетащите колонки сюда",
    "values": "Значения",
    "valueColumnsEmptyMessage": "Перетащите сюда для агрегации",
}

_CACHE = {"data": pd.DataFrame(), "last_modified": 0}


# --- ИНИЦИАЛИЗАЦИЯ И ИНДЕКСЫ БД ---
def init_db_indexes():
    if not os.path.exists(DB_NAME):
        return
    with sqlite3.connect(DB_NAME) as conn:
        cur = conn.cursor()
        cur.execute('CREATE INDEX IF NOT EXISTS idx_period ON medical_data ("Период");')
        cur.execute(
            'CREATE INDEX IF NOT EXISTS idx_dept ON medical_data ("Наименование отделения");'
        )
        cur.execute(
            'CREATE INDEX IF NOT EXISTS idx_profile ON medical_data ("Наименование профиля");'
        )
        cur.execute(
            'CREATE INDEX IF NOT EXISTS idx_mes ON medical_data ("Код Услуги");'
        )
        cur.execute(
            'CREATE INDEX IF NOT EXISTS idx_patient ON medical_data ("Номер ИБ");'
        )
        conn.commit()


init_db_indexes()

app = dash.Dash(
    __name__,
    external_stylesheets=[
        dbc.themes.BOOTSTRAP,
        "https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css",
        "https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap",
    ],
    external_scripts=[
        "https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js",
        "https://cdnjs.cloudflare.com/ajax/libs/vanilla-tilt/1.8.0/vanilla-tilt.min.js"
    ],
)
app.title = "Clinical Dashboard"


# --- НАСТРОЙКА КЭШИРОВАНИЯ ---
try:
    cache = Cache(
        app.server,
        config={
            "CACHE_TYPE": "RedisCache",
            "CACHE_REDIS_URL": "redis://127.0.0.1:6379/0",
            "CACHE_DEFAULT_TIMEOUT": 600,
        },
    )
    cache.set("test_key", "test_value")
except:
    logging.warning("Redis недоступен. Включен локальный файловый кэш FileSystemCache.")
    cache = Cache(
        app.server,
        config={
            "CACHE_TYPE": "FileSystemCache",
            "CACHE_DIR": "cache-directory",
            "CACHE_DEFAULT_TIMEOUT": 600,
        },
    )


# --- CSS СТИЛИ И КЛИЕНТСКИЙ JAVASCRIPT ---
app.index_string = r'''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            :root {
                --bg-color: #f4f7fe; --card-bg: #ffffff; --text-main: #1b2559;
                --text-muted: #a3aed1; --grid-color: #e9edf7;
                --shadow: 0px 18px 40px rgba(112, 144, 176, 0.12);
                --sidebar-shadow: 14px 17px 40px 4px rgba(112, 144, 176, 0.08);
                --primary: #4318FF; --primary-light: rgba(67, 24, 255, 0.1);
            }
            [data-theme="dark"] {
                --bg-color: #0b1437; --card-bg: #111c44; --text-main: #ffffff;
                --text-muted: #8f9bba; --grid-color: #1b254b;
                --primary-light: rgba(67, 24, 255, 0.2);
                --shadow: 0px 18px 40px rgba(0, 0, 0, 0.4);
                --sidebar-shadow: 14px 17px 40px 4px rgba(0, 0, 0, 0.5);
            }
            body { background-color: var(--bg-color); color: var(--text-main); font-family: 'Inter', sans-serif; transition: background-color 0.4s ease, color 0.4s ease; margin: 0; }
            .Select, .Select-value { background-color: transparent !important; }
            .Select-control { background-color: var(--card-bg) !important; border: 2px solid var(--grid-color) !important; border-radius: 16px !important; box-shadow: none !important; padding: 4px 8px !important; }
            .Select-control:hover, .is-focused > .Select-control { border-color: var(--primary) !important; box-shadow: 0 0 0 4px var(--primary-light) !important; }
            .Select-value-label, .Select-input > input, .Select-placeholder { color: var(--text-main) !important; font-weight: 500; }
            .Select-menu-outer { background-color: var(--card-bg) !important; border: 1px solid var(--grid-color) !important; border-radius: 16px !important; box-shadow: var(--shadow) !important; margin-top: 8px !important; padding: 8px !important; z-index: 9999 !important; }
            .Select-menu-outer input { background-color: var(--bg-color) !important; border: 1px solid var(--grid-color) !important; color: var(--text-main) !important; border-radius: 10px !important; padding: 10px 15px !important; }
            .Select-menu-outer label { color: var(--text-main) !important; font-weight: 500 !important; }
            .VirtualizedSelectOption { color: var(--text-main) !important; padding: 10px 15px !important; border-radius: 10px !important; margin-bottom: 2px !important; background-color: transparent !important; }
            .VirtualizedSelectFocusedOption, .VirtualizedSelectOption:hover { background-color: var(--primary-light) !important; color: var(--primary) !important; }
            .has-value.Select--multi .Select-value { background-color: var(--primary-light) !important; color: var(--primary) !important; border: 1px solid var(--grid-color) !important; border-radius: 12px !important; padding: 4px 10px !important; margin: 4px !important; font-weight: 600; }
            .has-value.Select--multi .Select-value-icon { border-right: 1px solid var(--grid-color) !important; color: var(--primary) !important; border-radius: 12px 0 0 12px !important; }
            .has-value.Select--multi .Select-value-icon:hover { background-color: var(--primary) !important; color: white !important; }
            ::-webkit-scrollbar { width: 8px; height: 8px; }
            ::-webkit-scrollbar-track { background: transparent; }
            ::-webkit-scrollbar-thumb { background: var(--grid-color); border-radius: 10px; }
            ::-webkit-scrollbar-thumb:hover { background: var(--text-muted); }
            .Select-clear-zone { color: var(--text-muted) !important; }
            .Select-arrow { border-color: var(--text-muted) transparent transparent !important; }
            .sql-editor { font-family: 'Courier New', Courier, monospace !important; background-color: #1e1e1e !important; color: #d4d4d4 !important; border-radius: 16px !important; border: 2px solid var(--grid-color) !important; padding: 15px !important; width: 100% !important; min-height: 140px !important; resize: vertical !important; font-size: 14px; }
            .sql-editor:focus { border-color: var(--primary) !important; outline: none !important; box-shadow: 0 0 0 4px var(--primary-light) !important; }
            .schema-badge { background-color: var(--bg-color); color: var(--text-main); padding: 8px 14px; border-radius: 10px; font-size: 13px; font-weight: 600; display: inline-block; margin: 4px; border: 1px solid var(--grid-color); }
            .nav-tabs { border-bottom: 2px solid var(--grid-color); margin-bottom: 25px; }
            .nav-tabs .nav-link { color: var(--text-muted); font-weight: 600; border: none; padding: 12px 25px; border-radius: 12px 12px 0 0; }
            .nav-tabs .nav-link:hover { color: var(--primary); background-color: var(--primary-light); }
            .nav-tabs .nav-link.active { color: var(--primary); background-color: transparent; border-bottom: 3px solid var(--primary); }
            @media print { .no-print { display: none !important; } body { background-color: white !important; } .card { box-shadow: none !important; border: 1px solid #ddd !important; } }

            [data-dash-is-loading="true"] {
                position: relative;
                pointer-events: none;
            }

            [data-dash-is-loading="true"] > * {
                opacity: 0 !important;
                transition: opacity 0.1s;
            }
            [data-dash-is-loading="true"]::before {
                content: '';
                position: absolute;
                top: 0; left: 0; right: 0; bottom: 0;
                background: linear-gradient(90deg, var(--card-bg) 25%, var(--grid-color) 50%, var(--card-bg) 75%);
                background-size: 200% 100%;
                animation: skeleton-loading 1.5s infinite linear;
                z-index: 100;
                border-radius: inherit;
            }
            @keyframes skeleton-loading {
                0% { background-position: 200% 0; }
                100% { background-position: -200% 0; }
            }
        </style>
    </head>
    <body>
        <div id="app-container">{%app_entry%}</div>
        <footer>
            {%config%} {%scripts%} {%renderer%}
            <script>
                // --- ШПИОН ЗА SHIFT ---
                window.isShiftPressed = false;
                document.addEventListener('keydown', function(e) { if(e.key === 'Shift') window.isShiftPressed = true; });
                document.addEventListener('keyup', function(e) { if(e.key === 'Shift') window.isShiftPressed = false; });
                window.addEventListener('blur', function() { window.isShiftPressed = false; });

                // --- АВТОСКОБКИ ДЛЯ SQL ---
                document.addEventListener('keydown', function(e) {
                    if (e.target && e.target.classList.contains('sql-editor')) {
                        const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype, "value").set;
                        const val = e.target.value;
                        const start = e.target.selectionStart;
                        const end = e.target.selectionEnd;
                        const pairs = { '(': ')', '[': ']', '{': '}', "'": "'", '"': '"' };
                        const closeChars = [')', ']', '}', "'", '"'];

                        if (closeChars.includes(e.key) && start === end && val[start] === e.key) {
                            e.preventDefault();
                            e.target.selectionStart = e.target.selectionEnd = start + 1;
                            return;
                        }

                        if (e.key === 'Tab') {
                            e.preventDefault();
                            const spaces = "    ";
                            nativeSetter.call(e.target, val.substring(0, start) + spaces + val.substring(end));
                            e.target.selectionStart = e.target.selectionEnd = start + spaces.length;
                            e.target.dispatchEvent(new Event('input', { bubbles: true }));
                            return;
                        }

                        if (pairs[e.key]) {
                            e.preventDefault();
                            const openChar = e.key;
                            const closeChar = pairs[openChar];
                            if (start !== end) {
                                const selected = val.substring(start, end);
                                nativeSetter.call(e.target, val.substring(0, start) + openChar + selected + closeChar + val.substring(end));
                                e.target.selectionStart = start + 1;
                                e.target.selectionEnd = end + 1;
                            } else {
                                nativeSetter.call(e.target, val.substring(0, start) + openChar + closeChar + val.substring(end));
                                e.target.selectionStart = e.target.selectionEnd = start + 1;
                            }
                            e.target.dispatchEvent(new Event('input', { bubbles: true }));
                        }
                    }
                });

                function animateCounter(element, endStr) {
                    let endVal = parseFloat(endStr.replace(/[^\d,-]/g, '').replace(',', '.'));
                    if (isNaN(endVal)) return;

                    element._isAnimating = true; // БЛОКИРУЕМ OBSERVER НА ВРЕМЯ АНИМАЦИИ
                    let startVal = 0;
                    let duration = 1000;
                    let startTime = null;

                    function step(currentTime) {
                        if (!startTime) startTime = currentTime;
                        let progress = Math.min((currentTime - startTime) / duration, 1);
                        let ease = 1 - Math.pow(1 - progress, 4);
                        let currentVal = startVal + (endVal - startVal) * ease;

                        if (endStr.includes('₽')) {
                            element.textContent = currentVal.toLocaleString('ru-RU', {minimumFractionDigits: 2, maximumFractionDigits: 2}) + ' ₽';
                        } else if (endStr.includes('чел.')) {
                            element.textContent = Math.floor(currentVal).toLocaleString('ru-RU') + ' чел.';
                        } else if (endStr.includes('ед.')) {
                            element.textContent = Math.floor(currentVal).toLocaleString('ru-RU') + ' ед.';
                        } else {
                            element.textContent = Math.floor(currentVal).toLocaleString('ru-RU');
                        }

                        if (progress < 1) {
                            window.requestAnimationFrame(step);
                        } else {
                            element.textContent = endStr; 
                            // Снимаем блокировку с легкой задержкой
                            setTimeout(() => { element._isAnimating = false; }, 50);
                        }
                    }
                    window.requestAnimationFrame(step);
                }

                const observer = new MutationObserver((mutations) => {
                    const cards = document.querySelectorAll('.tilt-card:not(.tilt-init)');
                    if (cards.length > 0 && window.VanillaTilt) {
                        VanillaTilt.init(cards, { max: 6, speed: 400, glare: true, "max-glare": 0.2, scale: 1.02 });
                        cards.forEach(c => c.classList.add('tilt-init'));
                    }

                    mutations.forEach(mutation => {
                        if (mutation.type === 'characterData' || mutation.type === 'childList') {
                            let target = mutation.target;
                            if (target.nodeType === 3) target = target.parentElement;
                            
                            if (target && target.classList && target.classList.contains('kpi-value')) {
                                let newVal = target.innerText;
                                if (target._dashValue !== newVal && !target._isAnimating && newVal !== "0") {
                                    target._dashValue = newVal;
                                    animateCounter(target, newVal);
                                }
                            }
                        }
                    });
                });
                
                observer.observe(document.body, { childList: true, subtree: true, characterData: true });
            </script>
        </footer>
    </body>
</html>
'''


# --- ФУНКЦИИ ОБРАБОТКИ ДАННЫХ И ВИЗУАЛА ---
def get_optimized_data() -> pd.DataFrame:
    if not os.path.exists(DB_NAME):
        return pd.DataFrame()
    current_mtime = os.path.getmtime(DB_NAME)
    if not _CACHE["data"].empty and _CACHE["last_modified"] == current_mtime:
        return _CACHE["data"]

    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql("SELECT * FROM medical_data", conn)
    conn.close()

    if df.empty:
        return df

    df.columns = (
        df.columns.str.replace(r"\n|\r", " ", regex=True)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    if "Период" in df.columns:
        df["dt"] = pd.to_datetime(df["Период"], errors="coerce")
        df = df.dropna(subset=["dt"])
        df["Year"] = df["dt"].dt.year.astype(int)
        df["Month_Num"] = df["dt"].dt.month.astype(int)
        df["Month_Name"] = df["Month_Num"].map(MONTHS_RU)
        df["Quarter_Num"] = df["dt"].dt.quarter.astype(int)
        df["Quarter_Name"] = df["Quarter_Num"].map(QUARTERS_RU)
        df["Month_Str"] = df["dt"].dt.strftime("%Y-%m")

    if "Сумма" in df.columns and df["Сумма"].dtype == object:
        df["Сумма"] = pd.to_numeric(
            df["Сумма"].astype(str).str.replace(",", ".").str.replace(" ", ""),
            errors="coerce",
        ).fillna(0)

    _CACHE["data"] = df
    _CACHE["last_modified"] = current_mtime
    return df


def apply_beautiful_layout(
    fig: go.Figure, theme: str, x_range=None, tickvals=None, ticktext=None
) -> go.Figure:
    if theme == "dark":
        text_color, grid_color, card_bg = "#ffffff", "#1b254b", "#111c44"
    else:
        text_color, grid_color, card_bg = "#1b2559", "#e9edf7", "#ffffff"

    xaxis_config = dict(
        showgrid=False,
        color=text_color,
        showline=True,
        linecolor=grid_color,
        linewidth=2,
        automargin=True,
    )
    if x_range is not None:
        xaxis_config["range"] = x_range
    if tickvals is not None and ticktext is not None:
        xaxis_config["tickmode"] = "array"
        xaxis_config["tickvals"] = tickvals
        xaxis_config["ticktext"] = ticktext

    fig.update_layout(
        font_family="Inter",
        separators=", ",
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=10, r=10, t=30, b=10),
        legend=dict(
            orientation="v",
            yanchor="middle",
            y=0.5,
            xanchor="left",
            x=1.02,
            font=dict(color=text_color, size=11),
        ),
        xaxis=xaxis_config,
        yaxis=dict(
            showgrid=True,
            gridcolor=grid_color,
            zeroline=False,
            color=text_color,
            automargin=True,
        ),
        hoverlabel=dict(
            bgcolor=card_bg,
            font_size=14,
            font_family="Inter",
            font_color=text_color,
            bordercolor=grid_color,
        ),
        hovermode="closest",
        barmode="group",
    )
    return fig


def create_kpi_card(
    title: str, id_value: str, icon_class: str, color_hex: str, bg_rgba: str
) -> dbc.Card:
    return dbc.Card(
        [
            dbc.CardBody(
                [
                    dbc.Row(
                        [
                            dbc.Col(
                                [
                                    html.P(
                                        title,
                                        style={
                                            "color": "var(--text-muted)",
                                            "fontWeight": "600",
                                            "fontSize": "13px",
                                            "marginBottom": "6px",
                                            "letterSpacing": "0.5px",
                                            "textTransform": "uppercase",
                                        },
                                    ),
                                    html.H3(
                                        id=id_value,
                                        className="kpi-value",
                                        style={
                                            "color": "var(--text-main)",
                                            "fontWeight": "800",
                                            "margin": "0",
                                            "fontSize": "26px",
                                            "transition": "color 0.3s",
                                        },
                                    ),
                                ],
                                width=8,
                                className="pe-0",
                            ),
                            dbc.Col(
                                [
                                    html.Div(
                                        html.I(
                                            className=icon_class,
                                            style={
                                                "color": color_hex,
                                                "fontSize": "22px",
                                            },
                                        ),
                                        style={
                                            "backgroundColor": bg_rgba,
                                            "width": "54px",
                                            "height": "54px",
                                            "borderRadius": "50%",
                                            "display": "flex",
                                            "alignItems": "center",
                                            "justifyContent": "center",
                                            "marginLeft": "auto",
                                        },
                                    )
                                ],
                                width=4,
                                className="ps-0",
                            ),
                        ],
                        className="align-items-center",
                    )
                ]
            )
        ],
        className="tilt-card",
        style={
            "backgroundColor": "var(--card-bg)",
            "border": "none",
            "borderRadius": "24px",
            "boxShadow": "var(--shadow)",
            "transition": "box-shadow 0.3s",
            "height": "100%",
            "position": "relative",
            "overflow": "hidden"
        },
    )


# --- ВЕРСТКА ИНТЕРФЕЙСА ---
app.layout = html.Div(
    [
        dcc.Store(id="theme-store", data="light"),
        dcc.Store(id="presets-store", storage_type="local", data={}),
        dcc.Store(id="filtered-click-data"),
        dcc.Store(id="filtered-selected-data"),
        html.Div(
            className="no-print",
            style={
                "position": "fixed",
                "top": 0,
                "left": 0,
                "bottom": 0,
                "width": "340px",
                "padding": "40px 30px",
                "backgroundColor": "var(--card-bg)",
                "boxShadow": "var(--sidebar-shadow)",
                "zIndex": 100,
                "overflowY": "auto",
            },
            children=[
                html.Div(
                    [
                        html.Img(
                            src=CUSTOM_ICON,
                            style={
                                "height": "45px",
                                "width": "auto",
                                "display": "block",
                                # "marginRight": "12px"
                            },
                        ),
                    ],
                    style={
                        "marginBottom": "35px",
                        "display": "flex",
                        "alignItems": "center",
                        "padding": "10px",
                    },
                ),
                html.Label(
                    "Поиск пациента",
                    style={
                        "color": "var(--text-main)",
                        "fontWeight": "700",
                        "fontSize": "13px",
                        "marginBottom": "8px",
                        "textTransform": "uppercase",
                    },
                ),
                dbc.Input(
                    id="f-patient",
                    placeholder="Введите номер ИБ и нажмите Enter...",
                    debounce=True,
                    className="mb-4",
                    style={
                        "borderRadius": "12px",
                        "border": "2px solid var(--grid-color)",
                        "padding": "10px",
                        "backgroundColor": "var(--bg-color)",
                        "color": "var(--text-main)",
                    },
                ),
                dbc.Checklist(
                    options=[
                        {"label": "Сравнить с прошлым годом (YoY)", "value": True}
                    ],
                    id="f-yoy",
                    value=[],
                    switch=True,
                    style={
                        "color": "var(--primary)",
                        "fontWeight": "600",
                        "marginBottom": "25px",
                    },
                ),
                html.Div(
                    [
                        html.Label(
                            "1. Показатель",
                            style={
                                "color": "var(--text-main)",
                                "fontWeight": "700",
                                "fontSize": "13px",
                                "marginBottom": "8px",
                                "textTransform": "uppercase",
                            },
                        ),
                        dcc.Dropdown(
                            id="f-metric",
                            options=[
                                {"label": "💰 Общая сумма (₽)", "value": "sum"},
                                {"label": "📋 Количество услуг", "value": "count_mes"},
                                {"label": "🧑 Пациенты", "value": "count_patients"},
                            ],
                            value="count_mes",
                            clearable=False,
                        ),
                        html.Label(
                            "2. Разрез линий",
                            style={
                                "color": "var(--text-main)",
                                "fontWeight": "700",
                                "fontSize": "13px",
                                "marginBottom": "8px",
                                "marginTop": "20px",
                                "textTransform": "uppercase",
                            },
                        ),
                        dcc.Dropdown(
                            id="f-group-by",
                            options=[
                                {
                                    "label": "🏥 По отделениям",
                                    "value": "Наименование отделения",
                                },
                                {"label": "📋 По кодам МЭС", "value": "Код Услуги"},
                                {
                                    "label": "🩺 По профилям",
                                    "value": "Наименование профиля",
                                },
                            ],
                            value="Наименование отделения",
                            clearable=False,
                        ),
                    ],
                    style={
                        "marginBottom": "25px",
                        "paddingBottom": "25px",
                        "borderBottom": "2px dashed var(--grid-color)",
                    },
                ),
                html.Label(
                    "ФИЛЬТРЫ ДАННЫХ",
                    style={
                        "color": "var(--text-muted)",
                        "fontWeight": "800",
                        "fontSize": "12px",
                        "marginBottom": "15px",
                        "letterSpacing": "1px",
                    },
                ),
                html.Div(
                    [
                        html.Label(
                            "Период (Год)",
                            style={
                                "color": "var(--text-main)",
                                "fontWeight": "600",
                                "fontSize": "14px",
                                "marginBottom": "8px",
                            },
                        ),
                        dcc.Dropdown(
                            id="f-year", multi=True, placeholder="Все года..."
                        ),
                    ]
                ),
                html.Div(
                    [
                        html.Label(
                            "Квартал",
                            style={
                                "color": "var(--text-main)",
                                "fontWeight": "600",
                                "fontSize": "14px",
                                "marginBottom": "8px",
                            },
                        ),
                        dcc.Dropdown(
                            id="f-quarter", multi=True, placeholder="Все кварталы..."
                        ),
                    ]
                ),
                html.Div(
                    [
                        html.Label(
                            "Месяц",
                            style={
                                "color": "var(--text-main)",
                                "fontWeight": "600",
                                "fontSize": "14px",
                                "marginBottom": "8px",
                            },
                        ),
                        dcc.Dropdown(
                            id="f-month", multi=True, placeholder="Все месяцы..."
                        ),
                    ]
                ),
                html.Div(
                    [
                        html.Label(
                            "Отделение",
                            style={
                                "color": "var(--text-main)",
                                "fontWeight": "600",
                                "fontSize": "14px",
                                "marginBottom": "8px",
                            },
                        ),
                        dcc.Dropdown(
                            id="f-dept", multi=True, placeholder="Все отделения..."
                        ),
                    ]
                ),
                html.Div(
                    [
                        html.Label(
                            "Профиль",
                            style={
                                "color": "var(--text-main)",
                                "fontWeight": "600",
                                "fontSize": "14px",
                                "marginBottom": "8px",
                            },
                        ),
                        dcc.Dropdown(
                            id="f-profile", multi=True, placeholder="Все профили..."
                        ),
                    ]
                ),
                html.Div(
                    [
                        html.Label(
                            "Код услуги",
                            style={
                                "color": "var(--text-main)",
                                "fontWeight": "600",
                                "fontSize": "14px",
                                "marginBottom": "8px",
                            },
                        ),
                        dcc.Dropdown(id="f-mes", multi=True, placeholder="Все коды..."),
                    ]
                ),
                dbc.Button([html.I(className="fas fa-check", style={"marginRight": "10px"}), "Применить фильтры"], id="btn-apply-filters", style={"width": "100%", "backgroundColor": "#01B574", "border": "none", "borderRadius": "16px", "padding": "14px", "fontWeight": "600", "marginTop": "25px", "boxShadow": "0px 10px 20px rgba(1, 181, 116, 0.2)"}),
                dbc.Button([html.I(className="fas fa-trash-alt", style={"marginRight": "10px"}), "Сбросить все фильтры"], id="btn-reset-all-filters", style={"width": "100%", "backgroundColor": "transparent", "border": "2px solid #E11D48", "color": "#E11D48", "borderRadius": "16px", "padding": "14px", "fontWeight": "600", "marginTop": "10px"}),
            ],
        ),
        html.Div(
            style={"marginLeft": "340px", "padding": "40px 50px", "minHeight": "100vh"},
            children=[
                dbc.Row(
                    [
                        dbc.Col(
                            html.H2(
                                [
                                    html.I(
                                        className="fas fa-heartbeat", 
                                        style={"marginRight": "12px", "color": "var(--primary)"}
                                    ),
                                    "Дашборд мониторинг"
                                ],
                                style={
                                    "fontWeight": "800",
                                    "color": "var(--text-main)",
                                    "letterSpacing": "-0.5px",
                                    "display": "flex",
                                    "alignItems": "center",
                                    "margin": 0
                                },
                            ),
                            width=6,
                        ),
                        dbc.Col(
                            html.Div(
                                [
                                    html.Button(
                                        html.I(className="fas fa-file-pdf"),
                                        id="btn-pdf",
                                        className="no-print",
                                        style={
                                            "background": "transparent",
                                            "border": "none",
                                            "fontSize": "22px",
                                            "color": "#E11D48",
                                            "cursor": "pointer",
                                            "marginRight": "20px",
                                        },
                                    ),
                                    html.Button(
                                        html.I(
                                            id="theme-icon", className="fas fa-moon"
                                        ),
                                        id="theme-toggle",
                                        className="no-print",
                                        style={
                                            "background": "transparent",
                                            "border": "none",
                                            "fontSize": "22px",
                                            "color": "var(--text-muted)",
                                            "cursor": "pointer",
                                            "marginRight": "25px",
                                        },
                                    ),
                                    html.Div(
                                        [
                                            html.Span(
                                                "Dr. Admin",
                                                style={
                                                    "fontWeight": "700",
                                                    "marginRight": "15px",
                                                    "color": "var(--text-main)",
                                                    "fontSize": "15px",
                                                },
                                            ),
                                            html.Img(
                                                src="https://ui-avatars.com/api/?name=Admin&background=4318FF&color=fff",
                                                style={
                                                    "borderRadius": "50%",
                                                    "width": "45px",
                                                },
                                            ),
                                        ],
                                        style={
                                            "display": "flex",
                                            "alignItems": "center",
                                            "backgroundColor": "var(--card-bg)",
                                            "border": "1px solid var(--grid-color)",
                                            "padding": "6px 20px",
                                            "borderRadius": "30px",
                                        },
                                    ),
                                ],
                                style={
                                    "textAlign": "right",
                                    "display": "flex",
                                    "alignItems": "center",
                                    "justifyContent": "flex-end",
                                },
                            ),
                            width=6,
                        ),
                    ],
                    style={"marginBottom": "30px", "marginTop": "10px"},
                ),
                dbc.Tabs(
                    id="main-tabs",
                    active_tab="tab-main",
                    children=[
                        dbc.Tab(
                            label="📈 Стандартный отчет",
                            tab_id="tab-main",
                            children=[
                                html.Div(
                                    style={"marginTop": "25px"},
                                    children=[
                                        dbc.Row(
                                            className="no-print",
                                            children=[
                                                dbc.Col(
                                                    create_kpi_card(
                                                        "ОБЩАЯ СУММА",
                                                        "kpi-sum",
                                                        "fas fa-wallet",
                                                        "#4318FF",
                                                        "rgba(67, 24, 255, 0.1)",
                                                    ),
                                                    xl=3,
                                                    lg=6,
                                                    md=6,
                                                    sm=12,
                                                    className="mb-4",
                                                ),
                                                dbc.Col(
                                                    create_kpi_card(
                                                        "УНИКАЛЬНЫХ ПАЦИЕНТОВ",
                                                        "kpi-patients",
                                                        "fas fa-user-injured",
                                                        "#FF7D00",
                                                        "rgba(255, 125, 0, 0.1)",
                                                    ),
                                                    xl=3,
                                                    lg=6,
                                                    md=6,
                                                    sm=12,
                                                    className="mb-4",
                                                ),
                                                dbc.Col(
                                                    create_kpi_card(
                                                        "ОКАЗАНО УСЛУГ (МЭС)",
                                                        "kpi-mes",
                                                        "fas fa-file-medical-alt",
                                                        "#01B574",
                                                        "rgba(1, 181, 116, 0.1)",
                                                    ),
                                                    xl=3,
                                                    lg=6,
                                                    md=6,
                                                    sm=12,
                                                    className="mb-4",
                                                ),
                                                dbc.Col(
                                                    create_kpi_card(
                                                        "АКТИВНЫХ ОТДЕЛЕНИЙ",
                                                        "kpi-depts",
                                                        "fas fa-hospital",
                                                        "#39B8FF",
                                                        "rgba(57, 184, 255, 0.1)",
                                                    ),
                                                    xl=3,
                                                    lg=6,
                                                    md=6,
                                                    sm=12,
                                                    className="mb-4",
                                                ),
                                            ],
                                        ),
                                        html.Div(
                                            id="pdf-export-container",
                                            children=[
                                                dbc.Row(
                                                    [
                                                        dbc.Col(
                                                            html.Div(
                                                                [
                                                                    html.H4(
                                                                        "Аналитика по времени",
                                                                        id="main-chart-title",
                                                                        style={
                                                                            "fontWeight": "800",
                                                                            "color": "var(--text-main)",
                                                                            "marginBottom": "15px",
                                                                        },
                                                                    ),
                                                                    dcc.Loading(
                                                                        type="dot",
                                                                        color="var(--primary)",
                                                                        children=[
                                                                            html.Div(
                                                                                id="smart-insights-container",
                                                                                style={
                                                                                    "marginBottom": "20px",
                                                                                    "padding": "16px 20px",
                                                                                    "backgroundColor": "var(--primary-light)",
                                                                                    "borderRadius": "14px",
                                                                                    "border": "1px solid var(--grid-color)",
                                                                                },
                                                                            ),
                                                                            dcc.Graph(
                                                                                id="main-line-chart",
                                                                                config={
                                                                                    "displayModeBar": False
                                                                                },
                                                                                style={
                                                                                    "height": "550px"
                                                                                },
                                                                            ),
                                                                        ],
                                                                    ),
                                                                ],
                                                                style={
                                                                    "backgroundColor": "var(--card-bg)",
                                                                    "borderRadius": "24px",
                                                                    "padding": "35px",
                                                                    "boxShadow": "var(--shadow)",
                                                                    "marginBottom": "25px",
                                                                },
                                                            ),
                                                            width=12,
                                                        )
                                                    ]
                                                ),
                                                dbc.Row(
                                                    className="mt-4 no-print",
                                                    children=[
                                                        dbc.Col(
                                                            html.Div(
                                                                [
                                                                    html.H4(
                                                                        [
                                                                            html.I(
                                                                                className="fas fa-terminal",
                                                                                style={
                                                                                    "marginRight": "12px",
                                                                                    "color": "var(--primary)",
                                                                                },
                                                                            ),
                                                                            "SQL-Песочница",
                                                                        ],
                                                                        style={
                                                                            "fontWeight": "800",
                                                                            "color": "var(--text-main)",
                                                                            "marginBottom": "15px",
                                                                        },
                                                                    ),
                                                                    html.P(
                                                                        "Напишите ваш SQL-запрос.",
                                                                        style={
                                                                            "color": "var(--text-muted)",
                                                                            "fontSize": "14px",
                                                                            "marginBottom": "20px",
                                                                        },
                                                                    ),
                                                                    html.Div(
                                                                        style={
                                                                            "position": "relative",
                                                                            "width": "100%",
                                                                        },
                                                                        children=[
                                                                            dcc.Textarea(
                                                                                id="sql-input",
                                                                                className="sql-editor",
                                                                                style={
                                                                                    "resize": "none",
                                                                                    "paddingRight": "50px",
                                                                                },
                                                                                placeholder="Введите ваш SQL запрос...",
                                                                                value="SELECT \n  [Наименование отделения], \n  COUNT(*) as [Количество услуг] \nFROM medical_data \nGROUP BY [Наименование отделения] \nORDER BY [Количество услуг] DESC \nLIMIT 7;",
                                                                            ),
                                                                            html.Button(
                                                                                html.I(
                                                                                    className="fas fa-expand"
                                                                                ),
                                                                                id="btn-expand-sql",
                                                                                title="Полноэкранный редактор",
                                                                                style={
                                                                                    "position": "absolute",
                                                                                    "top": "12px",
                                                                                    "right": "12px",
                                                                                    "background": "rgba(255, 255, 255, 0.05)",
                                                                                    "border": "none",
                                                                                    "borderRadius": "8px",
                                                                                    "padding": "8px",
                                                                                    "color": "#a3aed1",
                                                                                    "fontSize": "18px",
                                                                                    "cursor": "pointer",
                                                                                    "transition": "all 0.3s",
                                                                                },
                                                                            ),
                                                                        ],
                                                                    ),
                                                                    dbc.Button(
                                                                        [
                                                                            html.I(
                                                                                className="fas fa-play",
                                                                                style={
                                                                                    "marginRight": "10px"
                                                                                },
                                                                            ),
                                                                            "Выполнить SQL",
                                                                        ],
                                                                        id="btn-execute-sql",
                                                                        style={
                                                                            "backgroundColor": "var(--primary)",
                                                                            "border": "none",
                                                                            "borderRadius": "12px",
                                                                            "padding": "12px 24px",
                                                                            "fontWeight": "600",
                                                                            "marginTop": "15px",
                                                                        },
                                                                    ),
                                                                    html.Div(
                                                                        id="sql-error",
                                                                        style={
                                                                            "color": "#E11D48",
                                                                            "marginTop": "15px",
                                                                            "fontWeight": "600",
                                                                            "fontSize": "14px",
                                                                        },
                                                                    ),
                                                                ],
                                                                style={
                                                                    "backgroundColor": "var(--card-bg)",
                                                                    "borderRadius": "24px",
                                                                    "padding": "35px",
                                                                    "boxShadow": "var(--shadow)",
                                                                    "height": "100%",
                                                                },
                                                            ),
                                                            xl=8,
                                                            lg=12,
                                                            className="mb-4",
                                                        ),
                                                        dbc.Col(
                                                            html.Div(
                                                                [
                                                                    html.H5(
                                                                        [
                                                                            html.I(
                                                                                className="fas fa-database",
                                                                                style={
                                                                                    "marginRight": "10px",
                                                                                    "color": "#01B574",
                                                                                },
                                                                            ),
                                                                            "Схема Данных",
                                                                        ],
                                                                        style={
                                                                            "fontWeight": "800",
                                                                            "color": "var(--text-main)",
                                                                            "marginBottom": "25px",
                                                                        },
                                                                    ),
                                                                    html.Div(
                                                                        [
                                                                            html.Span(
                                                                                "Таблица:",
                                                                                style={
                                                                                    "color": "var(--text-muted)",
                                                                                    "fontSize": "13px",
                                                                                    "fontWeight": "600",
                                                                                    "textTransform": "uppercase",
                                                                                    "marginRight": "10px",
                                                                                },
                                                                            ),
                                                                            html.Span(
                                                                                "medical_data",
                                                                                style={
                                                                                    "color": "var(--primary)",
                                                                                    "fontWeight": "700",
                                                                                    "fontSize": "16px",
                                                                                    "backgroundColor": "var(--primary-light)",
                                                                                    "padding": "4px 10px",
                                                                                    "borderRadius": "8px",
                                                                                },
                                                                            ),
                                                                        ],
                                                                        style={
                                                                            "marginBottom": "25px"
                                                                        },
                                                                    ),
                                                                    html.P(
                                                                        "Доступные колонки:",
                                                                        style={
                                                                            "color": "var(--text-muted)",
                                                                            "fontSize": "13px",
                                                                            "fontWeight": "600",
                                                                            "textTransform": "uppercase",
                                                                            "marginBottom": "15px",
                                                                        },
                                                                    ),
                                                                    html.Div(
                                                                        [
                                                                            html.Span(
                                                                                "Период",
                                                                                className="schema-badge",
                                                                            ),
                                                                            html.Span(
                                                                                "Наименование отделения",
                                                                                className="schema-badge",
                                                                            ),
                                                                            html.Span(
                                                                                "Код Услуги",
                                                                                className="schema-badge",
                                                                            ),
                                                                            html.Span(
                                                                                "Наименование профиля",
                                                                                className="schema-badge",
                                                                            ),
                                                                            html.Span(
                                                                                "ИД пациента в версии счета",
                                                                                className="schema-badge",
                                                                            ),
                                                                            html.Span(
                                                                                "Сумма",
                                                                                className="schema-badge",
                                                                            ),
                                                                            html.Span(
                                                                                "Номер ИБ",
                                                                                className="schema-badge",
                                                                            ),
                                                                        ]
                                                                    ),
                                                                    html.Div(
                                                                        [
                                                                            html.I(
                                                                                className="fas fa-lightbulb",
                                                                                style={
                                                                                    "marginRight": "8px",
                                                                                    "color": "#FF7D00",
                                                                                },
                                                                            ),
                                                                            html.Span(
                                                                                "Если в названии колонки есть пробелы, оборачивайте её в квадратные скобки: ",
                                                                                style={
                                                                                    "color": "var(--text-muted)"
                                                                                },
                                                                            ),
                                                                            html.Code(
                                                                                "[Наименование отделения]",
                                                                                style={
                                                                                    "color": "var(--text-main)",
                                                                                    "fontWeight": "600",
                                                                                },
                                                                            ),
                                                                        ],
                                                                        style={
                                                                            "marginTop": "30px",
                                                                            "fontSize": "13px",
                                                                            "backgroundColor": "var(--bg-color)",
                                                                            "padding": "15px",
                                                                            "borderRadius": "12px",
                                                                            "border": "1px solid var(--grid-color)",
                                                                        },
                                                                    ),
                                                                ],
                                                                style={
                                                                    "backgroundColor": "var(--card-bg)",
                                                                    "borderRadius": "24px",
                                                                    "padding": "35px",
                                                                    "boxShadow": "var(--shadow)",
                                                                    "height": "100%",
                                                                },
                                                            ),
                                                            width=4,
                                                        ),
                                                    ],
                                                ),
                                            ],
                                        ),
                                    ],
                                )
                            ],
                        ),
                        dbc.Tab(
                            label="🧪 Улучшенная аналитика",
                            tab_id="tab-beta",
                            children=[
                                html.Div(
                                    style={"marginTop": "25px"},
                                    children=[
                                        dbc.Row(
                                            [
                                                dbc.Col(
                                                    html.Div(
                                                        [
                                                            html.Div(
                                                                [
                                                                    html.H4(
                                                                        [
                                                                            html.I(className="fas fa-th-large", style={"color": "#4318FF", "marginRight": "12px"}),
                                                                            "Дерево показателей (Treemap)",
                                                                        ],
                                                                        style={"fontWeight": "800", "color": "var(--text-main)", "margin": "0"},
                                                                    ),
                                                                    html.Div(
                                                                        [
                                                                            dbc.Checklist(
                                                                                options=[{"label": "Сетка (Одинаковые размеры)", "value": "equal"}],
                                                                                id="treemap-mode",
                                                                                value=[],
                                                                                switch=True,
                                                                                inline=True,
                                                                                style={"color": "var(--text-muted)", "fontWeight": "600", "marginRight": "20px"}
                                                                            ),
                                                                            html.Button(
                                                                                html.I(className="fas fa-expand-arrows-alt"),
                                                                                id="btn-expand-treemap",
                                                                                title="Растянуть в высоту",
                                                                                className="no-print",
                                                                                style={
                                                                                    "background": "rgba(67, 24, 255, 0.1)",
                                                                                    "border": "none",
                                                                                    "borderRadius": "8px", 
                                                                                    "padding": "8px 12px", 
                                                                                    "color": "var(--primary)", 
                                                                                    "cursor": "pointer", 
                                                                                    "fontSize": "16px"
                                                                                }
                                                                            )
                                                                        ], style={"display": "flex", "alignItems": "center"}
                                                                    )
                                                                ],
                                                                style={"display": "flex", "justifyContent": "space-between", "alignItems": "center", "marginBottom": "15px"}
                                                            ),
                                                            dcc.Loading(
                                                                type="dot",
                                                                color="#4318FF",
                                                                children=dcc.Graph(
                                                                    id="sunburst-chart",
                                                                    config={"displayModeBar": False},
                                                                    style={"height": "500px", "transition": "height 0.4s ease"}
                                                                ),
                                                            ),
                                                        ],
                                                        style={
                                                            "backgroundColor": "var(--card-bg)",
                                                            "borderRadius": "24px",
                                                            "padding": "35px",
                                                            "boxShadow": "var(--shadow)",
                                                            "marginBottom": "25px",
                                                        },
                                                    ),
                                                    width=12,
                                                )
                                            ]
                                        ),
                                        dbc.Row(
                                            [
                                                dbc.Col(
                                                    html.Div(
                                                        [
                                                            html.H4(
                                                                [
                                                                    html.I(
                                                                        className="fas fa-fire",
                                                                        style={
                                                                            "color": "#E11D48",
                                                                            "marginRight": "12px",
                                                                        },
                                                                    ),
                                                                    "Тепловая карта",
                                                                ],
                                                                style={
                                                                    "fontWeight": "800",
                                                                    "color": "var(--text-main)",
                                                                    "marginBottom": "15px",
                                                                },
                                                            ),
                                                            dcc.Loading(
                                                                type="dot",
                                                                color="#E11D48",
                                                                children=dcc.Graph(
                                                                    id="heatmap-chart",
                                                                    config={
                                                                        "displayModeBar": False
                                                                    },
                                                                ),
                                                            ),
                                                        ],
                                                        style={
                                                            "backgroundColor": "var(--card-bg)",
                                                            "borderRadius": "24px",
                                                            "padding": "35px",
                                                            "boxShadow": "var(--shadow)",
                                                            "marginBottom": "25px",
                                                        },
                                                    ),
                                                    width=12,
                                                )
                                            ]
                                        ),
                                        dbc.Row(
                                            [
                                                dbc.Col(
                                                    html.Div(
                                                        [
                                                            html.Div([
                                                                html.H4([html.I(className="fas fa-table", style={"color": "#01B574", "marginRight": "12px"}), "Сводная таблица (Конструктор)"], style={"fontWeight": "800", "margin": "0"}),
                                                                html.Div([
                                                                    html.Button([html.I(className="fas fa-times", style={"marginRight": "8px"}), "Сбросить фильтр"], id="btn-reset-crossfilter", style={"background": "rgba(225, 29, 72, 0.1)", "border": "1px solid #E11D48", "color": "#E11D48", "borderRadius": "8px", "padding": "6px 12px", "fontWeight": "600", "marginRight": "15px"}),
                                                                    html.Button(html.I(className="fas fa-file-csv"), id="btn-export-grid", title="Скачать таблицу (CSV)", className="no-print", style={"background": "transparent", "border": "none", "fontSize": "22px", "color": "#01B574", "cursor": "pointer"})
                                                                ], style={"display": "flex", "alignItems": "center"})
                                                            ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "center", "marginBottom": "15px"}),
                                                            html.Div(
                                                                id="crossfilter-msg",
                                                                style={
                                                                    "color": "#E11D48",
                                                                    "fontWeight": "700",
                                                                    "marginBottom": "15px",
                                                                },
                                                            ),
                                                            dcc.Loading(
                                                                type="dot",
                                                                color="#01B574",
                                                                children=html.Div(
                                                                    id="ag-grid-container",
                                                                    style={
                                                                        "width": "100%"
                                                                    },
                                                                ),
                                                            ),
                                                        ],
                                                        style={
                                                            "backgroundColor": "var(--card-bg)",
                                                            "borderRadius": "24px",
                                                            "padding": "35px",
                                                            "boxShadow": "var(--shadow)",
                                                        },
                                                    ),
                                                    width=12,
                                                )
                                            ]
                                        ),
                                    ],
                                )
                            ],
                        ),
                        dbc.Tab(
                            label="💼 Бизнес-аналитика",
                            tab_id="tab-abc",
                            children=[
                                html.Div(
                                    style={"marginTop": "25px"},
                                    children=[
                                        dbc.Row(
                                            [
                                                dbc.Col(
                                                    html.Div(
                                                        [
                                                            html.Div(
                                                                [
                                                                    html.H4(
                                                                        [
                                                                            html.I(
                                                                                className="fas fa-chart-pie",
                                                                                style={
                                                                                    "color": "#4318FF",
                                                                                    "marginRight": "12px",
                                                                                },
                                                                            ),
                                                                            "ABC-Анализ выручки",
                                                                        ],
                                                                        style={
                                                                            "fontWeight": "800",
                                                                            "margin": "0",
                                                                        },
                                                                    ),
                                                                    html.Button(
                                                                        html.I(
                                                                            className="fas fa-file-excel"
                                                                        ),
                                                                        id="btn-export-abc",
                                                                        title="Скачать цветной Excel",
                                                                        className="no-print",
                                                                        style={
                                                                            "background": "transparent",
                                                                            "border": "none",
                                                                            "fontSize": "22px",
                                                                            "color": "#01B574",
                                                                            "cursor": "pointer",
                                                                        },
                                                                    ),
                                                                ],
                                                                style={
                                                                    "display": "flex",
                                                                    "justifyContent": "space-between",
                                                                    "alignItems": "center",
                                                                    "marginBottom": "15px",
                                                                },
                                                            ),
                                                            html.P(
                                                                "Классификация по правилу Парето (Группа А: приносит 80% выручки, Группа В: 15%, Группа С: 5%).",
                                                                style={
                                                                    "color": "var(--text-muted)",
                                                                    "marginBottom": "20px",
                                                                },
                                                            ),
                                                            html.Label(
                                                                "Сгруппировать по:",
                                                                style={
                                                                    "color": "var(--text-main)",
                                                                    "fontWeight": "600",
                                                                    "fontSize": "13px",
                                                                },
                                                            ),
                                                            dcc.Dropdown(
                                                                id="abc-group-by",
                                                                options=[
                                                                    {
                                                                        "label": "📋 По кодам МЭС",
                                                                        "value": "Код Услуги",
                                                                    },
                                                                    {
                                                                        "label": "🏥 По отделениям",
                                                                        "value": "Наименование отделения",
                                                                    },
                                                                    {
                                                                        "label": "🩺 По профилям",
                                                                        "value": "Наименование профиля",
                                                                    },
                                                                ],
                                                                value="Код Услуги",
                                                                clearable=False,
                                                                style={
                                                                    "marginBottom": "25px",
                                                                    "width": "50%",
                                                                },
                                                            ),
                                                            dcc.Loading(
                                                                type="dot",
                                                                color="#4318FF",
                                                                children=html.Div(
                                                                    id="abc-grid-container",
                                                                    style={
                                                                        "width": "100%"
                                                                    },
                                                                ),
                                                            ),
                                                            dcc.Download(
                                                                id="download-abc-xlsx"
                                                            ),
                                                        ],
                                                        style={
                                                            "backgroundColor": "var(--card-bg)",
                                                            "borderRadius": "24px",
                                                            "padding": "35px",
                                                            "boxShadow": "var(--shadow)",
                                                        },
                                                    ),
                                                    width=12,
                                                )
                                            ]
                                        )
                                    ],
                                )
                            ],
                        ),
                    ],
                ),
            ],
        ),
        html.Div(
            id="selection-modal",
            style={"display": "none"},
            children=[
                html.Div(
                    style={
                        "position": "fixed", "bottom": "30px", "right": "30px", 
                        "backgroundColor": "var(--card-bg)", "padding": "25px", "borderRadius": "16px",
                        "boxShadow": "0px 20px 50px rgba(0,0,0,0.3)", "zIndex": "9999", "width": "350px",
                        "border": "2px solid var(--primary)", "color": "var(--text-main)"
                    },
                    children=[
                        html.Div([
                            html.I(className="fas fa-chart-pie", style={"color": "var(--primary)", "marginRight": "10px", "fontSize": "20px"}),
                            html.B("Аналитика по выбору", style={"fontSize": "18px"})
                        ], style={"marginBottom": "15px", "borderBottom": "1px solid var(--grid-color)", "paddingBottom": "10px"}),
                        
                        html.Div(id="selection-modal-content", style={"fontSize": "15px", "lineHeight": "1.8"}),
                        
                        html.Button(
                            "Закрыть", id="btn-close-modal", n_clicks=0,
                            style={
                                "marginTop": "20px", "width": "100%", "padding": "10px", "backgroundColor": "var(--primary)",
                                "color": "white", "border": "none", "borderRadius": "10px", "cursor": "pointer", "fontWeight": "bold"
                            }
                        )
                    ]
                )
            ]
        ),
        # --- МОДАЛЬНОЕ ОКНО: ДЕТАЛИЗАЦИЯ ---
        dbc.Modal(
            [
                dbc.ModalHeader(
                    dbc.ModalTitle(
                        "Расширенная информация",
                        id="modal-title",
                        style={"fontWeight": "800", "color": "var(--primary)"},
                    )
                ),
                dbc.ModalBody(id="modal-body", style={"padding": "25px"}),
                dbc.ModalFooter(
                    dbc.Button(
                        "Закрыть",
                        id="close-modal",
                        className="ms-auto",
                        style={"borderRadius": "12px"},
                    )
                ),
            ],
            id="drilldown-modal",
            is_open=False,
            size="lg",
            centered=True,
        ),

        dbc.Modal(
            [
                dbc.ModalHeader(
                    dbc.ModalTitle(
                        id="yoy-modal-title",
                        style={"fontWeight": "800", "color": "var(--primary)"},
                    )
                ),
                dbc.ModalBody(
                    id="yoy-modal-body", 
                    style={"padding": "25px", "backgroundColor": "var(--bg-color)"}
                ),
                dbc.ModalFooter(
                    dbc.Button(
                        "Закрыть",
                        id="close-yoy-modal",
                        className="ms-auto",
                        style={"borderRadius": "12px"},
                    )
                ),
            ],
            id="yoy-modal",
            is_open=False,
            size="lg",
            centered=True,
        ),

        # --- МОДАЛЬНОЕ ОКНО: SQL РЕДАКТОР ---
        dbc.Modal(
            [
                dbc.ModalHeader(
                    dbc.ModalTitle(
                        [
                            html.I(
                                className="fas fa-code", style={"marginRight": "10px"}
                            ),
                            "Продвинутый SQL-Редактор",
                        ],
                        style={"fontWeight": "800", "color": "var(--primary)"},
                    )
                ),
                dbc.ModalBody(
                    dbc.Row(
                        [
                            dbc.Col(
                                html.Div(
                                    [
                                        dcc.Textarea(
                                            id="sql-modal-input",
                                            className="sql-editor",
                                            style={
                                                "height": "480px",
                                                "resize": "none",
                                                "fontSize": "16px",
                                                "padding": "20px",
                                            },
                                        ),
                                        html.Div(
                                            id="sql-modal-linter",
                                            style={
                                                "marginTop": "8px",
                                                "fontSize": "13px",
                                                "fontWeight": "600",
                                                "minHeight": "20px",
                                            },
                                        ),
                                    ]
                                ),
                                width=8,
                            ),
                            dbc.Col(
                                html.Div(
                                    [
                                        html.H5(
                                            [
                                                html.I(
                                                    className="fas fa-table",
                                                    style={
                                                        "marginRight": "10px",
                                                        "color": "#01B574",
                                                    },
                                                ),
                                                "Схема: medical_data",
                                            ],
                                            style={
                                                "fontWeight": "800",
                                                "color": "var(--text-main)",
                                                "marginBottom": "20px",
                                            },
                                        ),
                                        html.P(
                                            "Доступные колонки:",
                                            style={
                                                "color": "var(--text-muted)",
                                                "fontSize": "13px",
                                                "fontWeight": "600",
                                                "textTransform": "uppercase",
                                                "marginBottom": "15px",
                                            },
                                        ),
                                        html.Div(
                                            [
                                                html.Span(
                                                    "Период",
                                                    className="schema-badge",
                                                    style={
                                                        "width": "100%",
                                                        "textAlign": "left",
                                                    },
                                                ),
                                                html.Span(
                                                    "Наименование отделения",
                                                    className="schema-badge",
                                                    style={
                                                        "width": "100%",
                                                        "textAlign": "left",
                                                    },
                                                ),
                                                html.Span(
                                                    "Код Услуги",
                                                    className="schema-badge",
                                                    style={
                                                        "width": "100%",
                                                        "textAlign": "left",
                                                    },
                                                ),
                                                html.Span(
                                                    "Наименование профиля",
                                                    className="schema-badge",
                                                    style={
                                                        "width": "100%",
                                                        "textAlign": "left",
                                                    },
                                                ),
                                                html.Span(
                                                    "ИД пациента в версии счета",
                                                    className="schema-badge",
                                                    style={
                                                        "width": "100%",
                                                        "textAlign": "left",
                                                    },
                                                ),
                                                html.Span(
                                                    "Сумма",
                                                    className="schema-badge",
                                                    style={
                                                        "width": "100%",
                                                        "textAlign": "left",
                                                    },
                                                ),
                                                html.Span(
                                                    "Номер ИБ",
                                                    className="schema-badge",
                                                    style={
                                                        "width": "100%",
                                                        "textAlign": "left",
                                                    },
                                                ),
                                            ]
                                        ),
                                        html.Div(
                                            [
                                                html.I(
                                                    className="fas fa-lightbulb",
                                                    style={
                                                        "marginRight": "8px",
                                                        "color": "#FF7D00",
                                                    },
                                                ),
                                                html.Span(
                                                    "Если в названии колонки есть пробелы, оборачивайте её в квадратные скобки: ",
                                                    style={
                                                        "color": "var(--text-muted)"
                                                    },
                                                ),
                                                html.Code(
                                                    "[Наименование отделения]",
                                                    style={
                                                        "color": "var(--text-main)",
                                                        "fontWeight": "600",
                                                    },
                                                ),
                                            ],
                                            style={
                                                "marginTop": "30px",
                                                "fontSize": "13px",
                                                "backgroundColor": "var(--bg-color)",
                                                "padding": "15px",
                                                "borderRadius": "12px",
                                                "border": "1px solid var(--grid-color)",
                                            },
                                        ),
                                    ],
                                    style={
                                        "backgroundColor": "var(--bg-color)",
                                        "borderRadius": "16px",
                                        "padding": "25px",
                                        "height": "100%",
                                        "border": "1px solid var(--grid-color)",
                                    },
                                ),
                                width=4,
                            ),
                        ]
                    )
                ),
                dbc.ModalFooter(
                    [
                        dbc.Button(
                            "Отмена",
                            id="btn-close-sql-modal",
                            style={
                                "backgroundColor": "transparent",
                                "border": "none",
                                "color": "var(--text-muted)",
                                "fontWeight": "600",
                                "marginRight": "15px",
                            },
                        ),
                        dbc.Button(
                            [
                                html.I(
                                    className="fas fa-save",
                                    style={"marginRight": "8px"},
                                ),
                                "Сохранить и Применить",
                            ],
                            id="btn-save-sql-modal",
                            style={
                                "backgroundColor": "var(--primary)",
                                "border": "none",
                                "borderRadius": "12px",
                                "fontWeight": "600",
                                "padding": "10px 20px",
                            },
                        ),
                    ]
                ),
            ],
            id="sql-editor-modal",
            is_open=False,
            size="xl",
            centered=True,
            backdrop="static",
        ),
        # --- МОДАЛЬНОЕ ОКНО: ПАЦИЕНТ ---
        dbc.Modal(
            [
                dbc.ModalHeader(
                    dbc.ModalTitle(
                        id="patient-modal-title",
                        style={
                            "fontWeight": "800",
                            "color": "var(--primary)",
                            "fontSize": "22px",
                        },
                    )
                ),
                dbc.ModalBody(
                    id="patient-modal-body",
                    style={"padding": "30px", "backgroundColor": "var(--bg-color)"},
                ),
                dbc.ModalFooter(
                    dbc.Button(
                        "Закрыть",
                        id="close-patient-modal",
                        className="ms-auto",
                        style={"borderRadius": "12px"},
                    )
                ),
            ],
            id="patient-modal",
            is_open=False,
            size="xl",
            centered=True,
        ),
    ]
)

def to_tuple(val):
    return tuple(val) if isinstance(val, list) else val

def clean_sqls(lst):
    return ", ".join([f"'{str(x).replace(chr(39), chr(39)+chr(39))}'" for x in lst])

def build_where_clause(years, quarters, months, depts, profiles, mes_list, patient=None):
    """Генератор SQL-условий для мгновенной DuckDB фильтрации (Zero-Copy)"""
    conditions = ["1=1"]
    if years: conditions.append(f'"Year" IN ({", ".join(map(str, years))})')
    if quarters: conditions.append(f'"Quarter_Name" IN ({clean_sqls(quarters)})')
    if months: conditions.append(f'"Month_Name" IN ({clean_sqls(months)})')
    if depts: conditions.append(f'"Наименование отделения" IN ({clean_sqls(depts)})')
    if profiles: conditions.append(f'"Наименование профиля" IN ({clean_sqls(profiles)})')
    if mes_list: conditions.append(f'"Код Услуги" IN ({clean_sqls(mes_list)})')
    if patient and str(patient).strip() != "":
        safe_pat = str(patient).strip().replace("'", "''")
        conditions.append(f'CAST("Номер ИБ" AS VARCHAR) ILIKE \'%{safe_pat}%\'')
    return " AND ".join(conditions)


@cache.memoize(timeout=600)
def build_tab_1_data(years, quarters, months, depts, profiles, mes_list, patient, metric, group_by_col, theme, yoy_toggle):
    df = get_optimized_data()
    if df.empty: return go.Figure(), "Аналитика", "0 ₽", "0", "0", "0", html.Div("Нет данных")

    where_clause = build_where_clause(years, quarters, months, depts, profiles, mes_list, patient)
    
    fig = go.Figure()
    x_range, tickvals, ticktext = None, [], []
    insight_html = html.Div()

    with duckdb.connect() as conn:
        conn.register('df', df)
        total_sum = conn.execute(f'SELECT SUM(Сумма) FROM (SELECT DISTINCT "Номер ИБ", "Код Услуги", dt, Сумма FROM df WHERE {where_clause})').fetchone()[0] or 0
        total_patients = conn.execute(f'SELECT COUNT(DISTINCT "ИД пациента в версии счета") FROM df WHERE {where_clause}').fetchone()[0] or 0
        total_mes = conn.execute(f'SELECT COUNT(*) FROM (SELECT DISTINCT "Номер ИБ", "Код Услуги", dt FROM df WHERE {where_clause})').fetchone()[0] or 0
        active_depts = conn.execute(f'SELECT COUNT(DISTINCT "Наименование отделения") FROM df WHERE {where_clause}').fetchone()[0] or 0

        dates_info = conn.execute(f"SELECT MIN(dt), MAX(dt) FROM df WHERE {where_clause}").fetchone()
        
        if dates_info[0] is not None:
            min_dt, max_dt = pd.Timestamp(dates_info[0]), pd.Timestamp(dates_info[1])
            x_range = [min_dt - pd.Timedelta(days=3), max_dt + pd.Timedelta(days=3)]
            
            unique_dates_df = conn.execute(f"SELECT DISTINCT dt FROM df WHERE {where_clause} AND dt IS NOT NULL ORDER BY dt").df()
            unique_dates = unique_dates_df['dt'].tolist()
            tickvals = unique_dates
            ticktext = [f"{MONTHS_RU[pd.Timestamp(d).month]} {pd.Timestamp(d).year}" for d in unique_dates]

            if group_by_col in df.columns:
                if metric == "sum": query = f'SELECT dt, "{group_by_col}", SUM(Сумма) as val FROM (SELECT DISTINCT "Номер ИБ", "Код Услуги", dt, "{group_by_col}", Сумма FROM df WHERE {where_clause}) GROUP BY dt, "{group_by_col}"'
                elif metric == "count_patients": query = f'SELECT dt, "{group_by_col}", COUNT(DISTINCT "ИД пациента в версии счета") as val FROM df WHERE {where_clause} GROUP BY dt, "{group_by_col}"'
                else: query = f'SELECT dt, "{group_by_col}", COUNT(*) as val FROM df WHERE {where_clause} GROUP BY dt, "{group_by_col}"'
                
                trend = conn.execute(query).df().sort_values('dt')
                top_query = f'SELECT "{group_by_col}" FROM trend GROUP BY "{group_by_col}" ORDER BY SUM(val) DESC LIMIT 5'
                top_groups = [row[0] for row in conn.execute(top_query).fetchall()]

                colors = [{"hex": "#4318FF", "rgba": "rgba(67, 24, 255, 0.15)"}, {"hex": "#FF7D00", "rgba": "rgba(255, 125, 0, 0.15)"}, {"hex": "#01B574", "rgba": "rgba(1, 181, 116, 0.15)"}, {"hex": "#39B8FF", "rgba": "rgba(57, 184, 255, 0.15)"}, {"hex": "#E11D48", "rgba": "rgba(225, 29, 72, 0.15)"}]

                for i, group_val in enumerate(top_groups):
                    g_data = trend[trend[group_by_col] == group_val]
                    c = colors[i % len(colors)]
                    custom_data_formatted = [f"STD|{group_val}"] * len(g_data)
                    
                    fig.add_trace(go.Scatter(
                        x=g_data['dt'], y=g_data['val'], name=str(group_val)[:32] + "...", mode='lines+markers',
                        line=dict(width=4, shape='spline', smoothing=1.3, color=c["hex"]), marker=dict(size=12, color=c["hex"]),
                        fill='tozeroy', fillcolor=c["rgba"], customdata=custom_data_formatted
                    ))
                    
            if 'trend' in locals() and not trend.empty:
                if metric == "sum":
                    monthly = conn.execute(f'SELECT Month_Str, SUM(Сумма) as val FROM (SELECT DISTINCT "Номер ИБ", "Код Услуги", Month_Str, Сумма FROM df WHERE {where_clause}) GROUP BY Month_Str ORDER BY Month_Str').df()
                    lbl, fmt = "Общая сумма", lambda x: f"{x:,.2f} ₽".replace(',',' ')
                elif metric == "count_patients":
                    monthly = conn.execute(f'SELECT Month_Str, COUNT(DISTINCT "ИД пациента в версии счета") as val FROM df WHERE {where_clause} GROUP BY Month_Str ORDER BY Month_Str').df()
                    lbl, fmt = "Уникальные пациенты", lambda x: f"{x:,.0f} чел.".replace(',',' ')
                else:
                    monthly = conn.execute(f'SELECT Month_Str, COUNT(*) as val FROM df WHERE {where_clause} GROUP BY Month_Str ORDER BY Month_Str').df()
                    lbl, fmt = "Оказано услуг", lambda x: f"{x:,.0f} ед.".replace(',',' ')

                if len(monthly) >= 2:
                    val_first, val_last = monthly.iloc[0]['val'], monthly.iloc[-1]['val']
                    diff = val_last - val_first
                    pct = (diff / val_first * 100) if val_first > 0 else 0
                    color, icon, sign = ("#01B574", "fa-arrow-up", "+") if diff >= 0 else ("#E11D48", "fa-arrow-down", "")
                    insight_html = html.Div([
                        html.Div([html.I(className="fas fa-robot", style={"marginRight": "8px", "color": "var(--primary)"}), html.B("Сводка:")], style={"marginBottom": "10px", "fontSize": "16px"}),
                        html.Span(f"Показатель «{lbl}» изменился с {fmt(val_first)} до {fmt(val_last)}. Разница: "),
                        html.Span([html.I(className=f"fas {icon}", style={"marginRight": "5px"}), f"{sign}{fmt(diff)} ({sign}{pct:.1f}%)"], style={"color": color, "fontWeight": "800", "backgroundColor": f"{color}20", "padding": "4px 8px", "borderRadius": "6px", "marginLeft": "8px"})
                    ])

    fig = apply_beautiful_layout(fig, theme, x_range, tickvals, ticktext)
    fig.update_xaxes(rangeslider_visible=False)
    fig.update_layout(clickmode='event+select')
    
    str_total_sum = f"{total_sum:,.2f} ₽".replace(",", " ").replace(".", ",")
    return fig, "Аналитика по времени", str_total_sum, str(total_patients), str(total_mes), str(active_depts), insight_html

@cache.memoize(timeout=600)
def build_tab_2_data(years, quarters, months, depts, profiles, mes_list, patient, metric, theme, tree_mode=None):
    df = get_optimized_data()
    if df.empty: return go.Figure(), go.Figure()

    where_clause = build_where_clause(years, quarters, months, depts, profiles, mes_list, patient)
    sunburst_fig, heatmap_fig = go.Figure(), go.Figure()

    with duckdb.connect() as conn:
        conn.register('df', df)
        
        has_data = conn.execute(f"SELECT COUNT(*) FROM df WHERE {where_clause}").fetchone()[0] > 0
        if has_data:
            if 'Month_Str' in df.columns and 'Наименование отделения' in df.columns:
                if metric == "sum": query_hm = f'SELECT "Наименование отделения", "Month_Str", SUM(Сумма) as val FROM (SELECT DISTINCT "Номер ИБ", "Код Услуги", "Month_Str", "Наименование отделения", Сумма FROM df WHERE {where_clause}) GROUP BY "Наименование отделения", "Month_Str"'
                elif metric == "count_patients": query_hm = f'SELECT "Наименование отделения", "Month_Str", COUNT(DISTINCT "ИД пациента в версии счета") as val FROM df WHERE {where_clause} GROUP BY "Наименование отделения", "Month_Str"'
                else: query_hm = f'SELECT "Наименование отделения", "Month_Str", COUNT(*) as val FROM df WHERE {where_clause} GROUP BY "Наименование отделения", "Month_Str"'
                    
                hm_data = conn.execute(query_hm).df()
                pivot = hm_data.pivot(index='Наименование отделения', columns='Month_Str', values='val').fillna(0)
                heatmap_fig.add_trace(go.Heatmap(z=pivot.values, x=pivot.columns, y=pivot.index, colorscale="Blues" if theme == "light" else "Viridis"))
                heatmap_fig = apply_beautiful_layout(heatmap_fig, theme)
                heatmap_fig.update_layout(height=max(400, (len(pivot.index) * 35) + 150), xaxis=dict(showgrid=False, type='category'), yaxis=dict(showgrid=False, automargin=True, tickfont=dict(size=11)))

            if set(['Наименование отделения', 'Код Услуги']).issubset(df.columns):

                query_tree = f"""
                    SELECT 
                        "Наименование отделения", 
                        "Код Услуги", 
                        SUM(Сумма) as Сумма, 
                        COUNT(*) as count_val 
                    FROM (
                        SELECT DISTINCT "Номер ИБ", "Код Услуги", "Наименование отделения", Сумма 
                        FROM df WHERE {where_clause}
                    ) 
                    GROUP BY "Наименование отделения", "Код Услуги" 
                """
                df_sun = conn.execute(query_tree).df()

                for col_name in ['Наименование отделения', 'Код Услуги']:
                    df_sun[col_name] = df_sun[col_name].fillna("Неизвестно")
                
                val_col = 'Сумма' if metric == "sum" else 'count_val'
                df_sun = df_sun.nlargest(150, val_col)
                
                is_equal_mode = tree_mode and "equal" in tree_mode
                
                if is_equal_mode:
                    df_sun['равный_вес'] = 1
                    target_values = 'равный_вес'
                else:
                    target_values = val_col

                sunburst_fig = px.treemap(
                    df_sun, 
                    path=['Наименование отделения', 'Код Услуги'], 
                    values=target_values, 
                    color='Сумма' if metric == 'sum' else None, 
                    color_continuous_scale='Blues'
                )

                if is_equal_mode:
                    ht_template = "<b>%{label}</b><br>Parent_id: %{parent}<br>Кол-во уникальных МЭС внутри: %{value} шт.<extra></extra>"
                elif metric == "sum":
                    ht_template = "<b>%{label}</b><br>Parent_id: %{parent}<br>Сумма: %{value:,.2f} ₽<extra></extra>"
                else:
                    ht_template = "<b>%{label}</b><br>Parent_id: %{parent}<br>Оказано услуг (без дублей): %{value:,.0f} шт.<extra></extra>"
                
                sunburst_fig.update_traces(
                    maxdepth=2, 
                    pathbar=dict(visible=True, textfont=dict(size=14, family="Inter", color="#1b2559")), 
                    root_color="#e9edf7", 
                    marker=dict(line=dict(width=2, color="#ffffff")),
                    texttemplate="<b>%{label}</b>",
                    hovertemplate=ht_template
                )
                
                sunburst_fig.update_layout(margin=dict(t=45, l=10, r=10, b=10), paper_bgcolor="rgba(0,0,0,0)")

    return sunburst_fig, heatmap_fig


@app.callback(
    [
        Output("f-year", "options"),
        Output("f-quarter", "options"),
        Output("f-month", "options"),
        Output("f-dept", "options"),
        Output("f-profile", "options"),
        Output("f-mes", "options"),
    ],
    [
        Input("f-year", "value"),
        Input("f-quarter", "value"),
        Input("f-month", "value"),
        Input("f-dept", "value"),
        Input("f-profile", "value"),
        Input("f-mes", "value"),
        Input("f-patient", "value"),
    ],
)
def update_smart_filters(years, quarters, months, depts, profiles, mes_list, patient):
    df = get_optimized_data()
    if df.empty:
        return [[], [], [], [], [], []]

    def get_mask_for(skip_col: str):
        mask = pd.Series(True, index=df.index)
        if skip_col != "Year" and years:
            mask &= df["Year"].isin(years)
        if skip_col != "Quarter" and quarters:
            mask &= df["Quarter_Name"].isin(quarters)
        if skip_col != "Month" and months:
            mask &= df["Month_Name"].isin(months)
        if skip_col != "Dept" and depts:
            mask &= df["Наименование отделения"].isin(depts)
        if skip_col != "Profile" and profiles:
            mask &= df["Наименование профиля"].isin(profiles)
        if skip_col != "MES" and mes_list:
            mask &= df["Код Услуги"].isin(mes_list)
        if patient and "Номер ИБ" in df.columns:
            mask &= df["Номер ИБ"].astype(str).str.contains(str(patient), case=False)
        return mask

    opts_year = (
        sorted(df.loc[get_mask_for("Year"), "Year"].unique())
        if "Year" in df.columns
        else []
    )
    opts_quarter = (
        [
            {"label": row["Quarter_Name"], "value": row["Quarter_Name"]}
            for _, row in df.loc[get_mask_for("Quarter")]
            .drop_duplicates(["Quarter_Num", "Quarter_Name"])
            .sort_values("Quarter_Num")
            .iterrows()
        ]
        if "Quarter_Num" in df.columns
        else []
    )
    opts_month = (
        [
            {"label": row["Month_Name"], "value": row["Month_Name"]}
            for _, row in df.loc[get_mask_for("Month")]
            .drop_duplicates(["Month_Num", "Month_Name"])
            .sort_values("Month_Num")
            .iterrows()
        ]
        if "Month_Num" in df.columns
        else []
    )
    opts_dept = (
        sorted(df.loc[get_mask_for("Dept"), "Наименование отделения"].dropna().unique())
        if "Наименование отделения" in df.columns
        else []
    )
    opts_profile = (
        sorted(
            df.loc[get_mask_for("Profile"), "Наименование профиля"].dropna().unique()
        )
        if "Наименование профиля" in df.columns
        else []
    )
    opts_mes = (
        sorted(df.loc[get_mask_for("MES"), "Код Услуги"].dropna().unique())
        if "Код Услуги" in df.columns
        else []
    )

    return [opts_year, opts_quarter, opts_month, opts_dept, opts_profile, opts_mes]


# 🚀 1. Вкладка 1 (График + KPI)
@app.callback(
    [
        Output("main-line-chart", "figure"), Output("main-chart-title", "children"), Output("kpi-sum", "children"),
        Output("kpi-patients", "children"), Output("kpi-mes", "children"), Output("kpi-depts", "children"),
        Output("smart-insights-container", "children")
    ],
    [
        Input("btn-apply-filters", "n_clicks"),
        Input("btn-reset-all-filters", "n_clicks"),
        Input("main-tabs", "active_tab"),
        Input("theme-store", "data")
    ],
    [
        State("f-year", "value"), State("f-quarter", "value"), State("f-month", "value"), State("f-dept", "value"),
        State("f-profile", "value"), State("f-mes", "value"), State("f-patient", "value"), State("f-metric", "value"), 
        State("f-group-by", "value"), State("f-yoy", "value")
    ]
)
def update_tab_1_router(n_clicks_apply, n_clicks_reset, active_tab, theme, years, quarters, months, depts, profiles, mes_list, patient, metric, group_by_col, yoy_toggle):
    if active_tab != "tab-main":
        return dash.no_update, dash.no_update, dash.no_update, dash.no_update, dash.no_update, dash.no_update, dash.no_update

    ctx = dash.callback_context
    trig = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else None
    
    if trig == "btn-reset-all-filters":
        years, quarters, months, depts, profiles, mes_list, patient, yoy_toggle = None, None, None, None, None, None, "", []
    
    return build_tab_1_data(to_tuple(years), to_tuple(quarters), to_tuple(months), to_tuple(depts), to_tuple(profiles), to_tuple(mes_list), patient, metric, group_by_col, theme, to_tuple(yoy_toggle))


# 🚀 2. Вкладка 2 (Дерево + Хитмап)
@app.callback(
    [Output("sunburst-chart", "figure"), Output("heatmap-chart", "figure")],
    [
        Input("btn-apply-filters", "n_clicks"),
        Input("btn-reset-all-filters", "n_clicks"),
        Input("main-tabs", "active_tab"),
        Input("theme-store", "data"),
        Input("treemap-mode", "value")
    ],
    [
        State("f-year", "value"), State("f-quarter", "value"), State("f-month", "value"), State("f-dept", "value"),
        State("f-profile", "value"), State("f-mes", "value"), State("f-patient", "value"), State("f-metric", "value")
    ]
)
def update_tab_2_router(n_clicks_apply, n_clicks_reset, active_tab, theme, tree_mode, years, quarters, months, depts, profiles, mes_list, patient, metric):
    if active_tab != "tab-beta":
        return dash.no_update, dash.no_update
        
    ctx = dash.callback_context
    trig = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else None
    
    if trig == "btn-reset-all-filters":
        years, quarters, months, depts, profiles, mes_list, patient = None, None, None, None, None, None, ""
        
    return build_tab_2_data(to_tuple(years), to_tuple(quarters), to_tuple(months), to_tuple(depts), to_tuple(profiles), to_tuple(mes_list), patient, metric, theme, to_tuple(tree_mode))

# 🚀 3. Таблица
@app.callback(
    [Output("ag-grid-container", "children"), Output("crossfilter-msg", "children")],
    [
        Input("btn-apply-filters", "n_clicks"),
        Input("btn-reset-all-filters", "n_clicks"),
        Input("heatmap-chart", "clickData"),
        Input("main-tabs", "active_tab"),
        Input("theme-store", "data")
    ],
    [
        State("f-year", "value"), State("f-quarter", "value"), State("f-month", "value"), State("f-dept", "value"),
        State("f-profile", "value"), State("f-mes", "value"), State("f-patient", "value")
    ]
)

def update_table_only(n_clicks_apply, n_clicks_reset, heatmap_click, active_tab, theme, years, quarters, months, depts, profiles, mes_list, patient):

    if active_tab != "tab-beta":
        return dash.no_update, dash.no_update
        
    ctx = dash.callback_context
    trig = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else None
    
    if trig == "btn-reset-all-filters":
        years, quarters, months, depts, profiles, mes_list, patient, heatmap_click = None, None, None, None, None, None, "", None
        
    df = get_optimized_data()
    cf_msg = ""
    where_clause = build_where_clause(years, quarters, months, depts, profiles, mes_list, patient)

    if heatmap_click:
        point = heatmap_click['points'][0]
        clicked_month, clicked_dept = str(point['x'])[:7], str(point['y'])
        safe_dept = clicked_dept.replace("'", "''")
        where_clause += f" AND Month_Str = '{clicked_month}' AND \"Наименование отделения\" = '{safe_dept}'"
        cf_msg = f"🔍 Отфильтровано по: {clicked_dept} ({clicked_month})"

    with duckdb.connect() as conn:
        conn.register('df', df)
        
        has_data = conn.execute(f"SELECT COUNT(*) FROM df WHERE {where_clause}").fetchone()[0] > 0
        if not has_data: 
            return html.Div("Нет данных для отображения", style={"padding": "20px", "fontWeight": "bold"}), cf_msg

        agg_cols = ['Наименование отделения', 'Наименование профиля', 'Код Услуги', 'Номер ИБ']
        agg_cols = [c for c in agg_cols if c in df.columns]
        
        cols_str = ', '.join([f'"{c}"' for c in agg_cols])
        
        if 'Сумма' in df.columns:
            query = f'SELECT {cols_str}, COUNT(*) as Кол_во_услуг, ROUND(SUM(Сумма), 2) as Сумма FROM df WHERE {where_clause} GROUP BY {cols_str} ORDER BY Сумма DESC LIMIT 1000'
        else:
            query = f'SELECT {cols_str}, COUNT(*) as Кол_во_услуг FROM df WHERE {where_clause} GROUP BY {cols_str} ORDER BY Кол_во_услуг DESC LIMIT 1000'
            
        grid_df = conn.execute(query).df()

    if grid_df.empty:
        return html.Div("Нет данных для отображения", style={"padding": "20px", "fontWeight": "bold"}), cf_msg

    column_defs = []
    for col in grid_df.columns:
        col_def = {"field": col, "sortable": True, "enableRowGroup": True, "enablePivot": True}
        if col in ['Кол_во_услуг', 'Сумма']:
            col_def.update({"filter": "agNumberColumnFilter", "enableValue": True})
            if col == 'Сумма': col_def["aggFunc"] = "sum"
        else:
            col_def["filter"] = "agSetColumnFilter"
        column_defs.append(col_def)

    ag_grid = dag.AgGrid(
        id="interactive-grid",
        rowData=grid_df.to_dict("records"), columnDefs=column_defs, 
        defaultColDef={"flex": 1, "minWidth": 150, "floatingFilter": True}, 
        className="ag-theme-alpine" if theme == "light" else "ag-theme-alpine-dark", 
        enableEnterpriseModules=True, 
        dashGridOptions={"localeText": AG_GRID_LOCALE_RU, "pagination": True, "paginationPageSize": 15, "rowGroupPanelShow": "always", "sideBar": True}, 
        style={"height": "650px", "width": "100%", "borderRadius": "12px"}
    )

    return ag_grid, cf_msg


# 🚀 4. Вкладка 3 (ABC Анализ через DuckDB)
@app.callback(
    [Output("abc-grid-container", "children"), Output("download-abc-xlsx", "data")],
    [
        Input("btn-apply-filters", "n_clicks"),
        Input("btn-reset-all-filters", "n_clicks"),
        Input("abc-group-by", "value"),
        Input("btn-export-abc", "n_clicks"),
        Input("main-tabs", "active_tab"),
        Input("theme-store", "data")
    ],
    [
        State("f-year", "value"), State("f-quarter", "value"), State("f-month", "value"),
        State("f-dept", "value"), State("f-profile", "value"), State("f-mes", "value")
    ]
)
def update_abc_analysis(n_clicks_apply, n_clicks_reset, group_by, export_clicks, active_tab, theme, years, quarters, months, depts, profiles, mes_list):
    if active_tab != "tab-abc":
        return dash.no_update, dash.no_update
        
    ctx = dash.callback_context
    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else None

    if triggered_id == "btn-reset-all-filters":
        years, quarters, months, depts, profiles, mes_list = None, None, None, None, None, None

    df = get_optimized_data()
    if df.empty or 'Сумма' not in df.columns or group_by not in df.columns:
        return html.Div("Нет данных для ABC анализа"), dash.no_update

    where_clause = build_where_clause(years, quarters, months, depts, profiles, mes_list)
    
    with duckdb.connect() as conn:
        conn.register('df', df)
        
        has_data = conn.execute(f"SELECT COUNT(*) FROM df WHERE {where_clause}").fetchone()[0] > 0
        if not has_data:
            return html.Div("Нет данных для отображения", style={"padding": "20px", "fontWeight": "bold"}), dash.no_update

        query_abc = f"""
            SELECT "{group_by}", SUM(Сумма) as Сумма, COUNT(DISTINCT "Номер ИБ") as Кол_во_услуг
            FROM (SELECT DISTINCT "Номер ИБ", "Код Услуги", "{group_by}", Сумма FROM df WHERE {where_clause})
            GROUP BY "{group_by}"
            ORDER BY Сумма DESC
        """
        df_abc = conn.execute(query_abc).df()

    total_sum = df_abc['Сумма'].sum()
    if total_sum > 0:
        df_abc['Доля_выручки'] = df_abc['Сумма'] / total_sum
        df_abc['Накопленная_доля'] = df_abc['Доля_выручки'].cumsum()
        df_abc['Группа ABC'] = df_abc['Накопленная_доля'].apply(lambda pct: 'A' if pct <= 0.80 else ('B' if pct <= 0.95 else 'C'))
        df_abc['Сумма'] = df_abc['Сумма'].round(2)
        df_abc['Доля_выручки'] = (df_abc['Доля_выручки'] * 100).round(2).astype(str) + "%"
        df_abc['Накопленная_доля'] = (df_abc['Накопленная_доля'] * 100).round(2).astype(str) + "%"
    else:
        df_abc['Доля_выручки'] = "0%"
        df_abc['Накопленная_доля'] = "0%"
        df_abc['Группа ABC'] = "C"

    if triggered_id == "btn-export-abc":
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_abc.to_excel(writer, index=False, sheet_name='ABC Analysis')
            worksheet = writer.sheets['ABC Analysis']
            fill_a = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            fill_b = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
            fill_c = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            abc_col_idx = df_abc.columns.get_loc('Группа ABC') + 1

            for row_idx in range(2, len(df_abc) + 2):
                group_val = worksheet.cell(row=row_idx, column=abc_col_idx).value
                fill = fill_a if group_val == 'A' else (fill_b if group_val == 'B' else fill_c)
                for col_idx in range(1, len(df_abc.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill

        output.seek(0)
        return dash.no_update, dcc.send_bytes(output.getvalue(), "ABC_Analysis_Export.xlsx")

    if df_abc.empty:
        return html.Div("Нет данных", style={"padding": "20px", "fontWeight": "bold"}), dash.no_update

    column_defs = []
    for col in df_abc.columns:
        col_def = {"field": col, "sortable": True, "filter": True}
        if col in ['Сумма', 'Кол_во_услуг']: col_def["filter"] = "agNumberColumnFilter"
        else: col_def["filter"] = "agSetColumnFilter"
        column_defs.append(col_def)

    row_class_rules = {"abc-a": "data['Группа ABC'] == 'A'", "abc-b": "data['Группа ABC'] == 'B'", "abc-c": "data['Группа ABC'] == 'C'"}
    grid_theme = "ag-theme-alpine" if theme == "light" else "ag-theme-alpine-dark"

    ag_grid = dag.AgGrid(
        rowData=df_abc.to_dict("records"), columnDefs=column_defs, defaultColDef={"flex": 1, "minWidth": 120, "floatingFilter": True},
        className=grid_theme, rowClassRules=row_class_rules, enableEnterpriseModules=True,
        dashGridOptions={"localeText": AG_GRID_LOCALE_RU, "pagination": True, "paginationPageSize": 20},
        style={"height": "600px", "width": "100%", "borderRadius": "12px", "overflow": "hidden"}
    )
    return ag_grid, dash.no_update


# --- ДЕТАЛИЗАЦИЯ (ИСПРАВЛЕННОЕ МОДАЛЬНОЕ ОКНО) ---
@app.callback(
    [
        Output("drilldown-modal", "is_open"),
        Output("modal-title", "children"),
        Output("modal-body", "children"),
    ],
    [Input("filtered-click-data", "data"), Input("close-modal", "n_clicks")],
    [
        State("drilldown-modal", "is_open"),
        State("f-group-by", "value"),
        State("f-metric", "value"),
        State("f-year", "value"),
        State("f-quarter", "value"),
        State("f-month", "value"),
        State("f-dept", "value"),
        State("f-profile", "value"),
        State("f-mes", "value"),
        State("f-patient", "value"),
    ],
)
def drilldown_modal(clickData, close_clicks, is_open, group_by_col, metric, years, quarters, months, depts, profiles, mes_list, patient):
    ctx = dash.callback_context
    if not ctx.triggered:
        return is_open, dash.no_update, dash.no_update

    trigger_id = ctx.triggered[0]["prop_id"].split(".")[0]
    if trigger_id == "close-modal":
        return False, "", ""

    if clickData and trigger_id == "filtered-click-data":
        point = clickData["points"][0]
        custom_info = str(point.get("customdata", ""))

        if custom_info.startswith("SQL"):
            return (
                True,
                "Информация",
                html.Div("Детализация доступна только в стандартном режиме."),
            )

        clicked_date = point["x"]
        clicked_group = custom_info.split("|")[1] if "|" in custom_info else custom_info

        df = get_optimized_data()

        mask = pd.Series(True, index=df.index)
        if years:
            mask &= df["Year"].isin(years)
        if quarters:
            mask &= df["Quarter_Name"].isin(quarters)
        if months:
            mask &= df["Month_Name"].isin(months)
        if depts:
            mask &= df["Наименование отделения"].isin(depts)
        if profiles:
            mask &= df["Наименование профиля"].isin(profiles)
        if mes_list:
            mask &= df["Код Услуги"].isin(mes_list)
        if patient and "Номер ИБ" in df.columns:
            mask &= df["Номер ИБ"].astype(str).str.contains(str(patient), case=False)

        clicked_date_str = clicked_date.split(" ")[0]
        mask &= df["dt"].dt.strftime("%Y-%m-%d") == clicked_date_str
        mask &= df[group_by_col].astype(str).str.strip() == str(clicked_group).strip()

        df_filtered = df[mask]

        if group_by_col == "Наименование отделения":
            drill_col, drill_label = "Код Услуги", "МЭС"
        elif group_by_col == "Код Услуги":
            drill_col, drill_label = "Наименование отделения", "отделение"
        else:
            drill_col, drill_label = "Наименование отделения", "отделение"

        if metric == "sum":
            df_unique = df_filtered.drop_duplicates(subset=["Номер ИБ", "Код Услуги", drill_col, "Сумма"])
            total_val = df_unique["Сумма"].sum()
            fmt_val = f"{total_val:,.2f} ₽".replace(",", " ").replace(".", ",")
            breakdown = (
                df_unique.groupby(drill_col, observed=True)["Сумма"]
                .sum()
                .reset_index()
            )
            val_col, suffix = "Сумма", " ₽"
        elif metric == "count_patients":
            total_val = df_filtered["ИД пациента в версии счета"].nunique()
            fmt_val = f"{total_val} чел."
            breakdown = (
                df_filtered.groupby(drill_col, observed=True)[
                    "ИД пациента в версии счета"
                ]
                .nunique()
                .reset_index(name="val")
            )
            val_col, suffix = "val", " чел."
        else:
            total_val = len(df_filtered)
            fmt_val = f"{total_val} ед."
            breakdown = (
                df_filtered.groupby(drill_col, observed=True)
                .size()
                .reset_index(name="val")
            )
            val_col, suffix = "val", " ед."

        top_n = breakdown.sort_values(val_col, ascending=False).head(10)
        leader_name = top_n.iloc[0][drill_col] if not top_n.empty else "Н/Д"
        leader_val = top_n.iloc[0][val_col] if not top_n.empty else 0
        fmt_leader_val = (
            f"{leader_val:,.2f}".replace(",", " ").replace(".", ",")
            if metric == "sum"
            else f"{leader_val}"
        )

        summary_text = html.Div(
            [
                html.Span("В выбранный период по объекту "),
                html.B(f"«{clicked_group}»", style={"color": "var(--primary)"}),
                html.Span(" зафиксировано "),
                html.B(fmt_val),
                html.Span(
                    f". Из них основной объем (Топ-1) приходится на {drill_label} "
                ),
                html.B(f"«{leader_name}»"),
                html.Span(f", что составляет "),
                html.B(f"{fmt_leader_val}{suffix}"),
                html.Span("."),
            ],
            style={
                "backgroundColor": "var(--bg-color)",
                "padding": "15px 20px",
                "borderRadius": "12px",
                "fontSize": "15px",
                "lineHeight": "1.6",
                "color": "var(--text-main)",
                "marginBottom": "25px",
                "borderLeft": "5px solid var(--primary)",
            },
        )

        top_n_plot = top_n.sort_values(val_col, ascending=True)

        def shorten_label(label, max_len=40):
            return str(label) if len(str(label)) <= max_len else str(label)[:37] + "..."

        short_labels = top_n_plot[drill_col].apply(shorten_label)

        text_template = "%{x:,.2f} ₽" if metric == "sum" else "%{x:,.0f}"
        ht_template = (
            "<b>%{customdata}</b><br>Значение: %{x:,.2f} ₽<extra></extra>"
            if metric == "sum"
            else "<b>%{customdata}</b><br>Значение: %{x:,.0f}<extra></extra>"
        )

        fig_modal = go.Figure(
            go.Bar(
                x=top_n_plot[val_col],
                y=short_labels,
                customdata=top_n_plot[drill_col],
                orientation="h",
                marker_color="var(--primary)",
                texttemplate=text_template,
                textposition="outside",
                cliponaxis=False,
                hovertemplate=ht_template,
            )
        )

        max_x_val = top_n_plot[val_col].max() if not top_n_plot.empty else 0
        fig_modal.update_layout(
            margin=dict(r=20, t=10, b=10),
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(
                showgrid=True,
                gridcolor="var(--grid-color)",
                range=[0, (max_x_val * 1.15) if max_x_val > 0 else 10],
            ),
            yaxis=dict(type="category", automargin=True),
        )

        modal_body = html.Div(
            [
                summary_text,
                html.P(
                    f"Распределение по разрезу «{drill_label}»:",
                    style={
                        "fontWeight": "700",
                        "fontSize": "14px",
                        "marginBottom": "10px",
                    },
                ),
                dcc.Graph(
                    figure=fig_modal,
                    config={"displayModeBar": False},
                    style={"height": "380px"},
                ),
            ]
        )
        return True, f"Детальный анализ: {clicked_group}", modal_body
    return is_open, dash.no_update, dash.no_update


# --- КАРТОЧКА ПАЦИЕНТА ---
@app.callback(
    [
        Output("patient-modal", "is_open"),
        Output("patient-modal-title", "children"),
        Output("patient-modal-body", "children"),
    ],
    [Input("f-patient", "value"), Input("close-patient-modal", "n_clicks")],
    [State("patient-modal", "is_open"), State("theme-store", "data")],
    prevent_initial_call=True,
)
def toggle_patient_modal(patient_id, close_clicks, is_open, theme):
    ctx = dash.callback_context
    if not ctx.triggered:
        return is_open, dash.no_update, dash.no_update

    trigger_id = ctx.triggered[0]["prop_id"].split(".")[0]
    if trigger_id == "close-patient-modal":
        return False, dash.no_update, dash.no_update

    if trigger_id == "f-patient" and patient_id and str(patient_id).strip() != "":
        df = get_optimized_data()
        if df.empty or "Номер ИБ" not in df.columns:
            return is_open, dash.no_update, dash.no_update

        pat_df = df[
            df["Номер ИБ"].astype(str).str.contains(str(patient_id), case=False)
        ].copy()

        dedup_cols = ["Номер ИБ", "Код Услуги", "Сумма"]
        if "dt" in pat_df.columns:
            dedup_cols.append("dt")
        pat_df = pat_df.drop_duplicates(subset=dedup_cols)
        
        if pat_df.empty:
            return is_open, dash.no_update, dash.no_update

        total_spent = pat_df["Сумма"].sum() if "Сумма" in pat_df.columns else 0
        total_visits = len(pat_df)
        depts_visited = (
            pat_df["Наименование отделения"].nunique()
            if "Наименование отделения" in pat_df.columns
            else 0
        )

        if "dt" in pat_df.columns and not pat_df["dt"].isnull().all():
            min_dt, max_dt = pat_df["dt"].min(), pat_df["dt"].max()
            first_date = f"{MONTHS_RU.get(min_dt.month, '')} {min_dt.year}"
            last_date = f"{MONTHS_RU.get(max_dt.month, '')} {max_dt.year}"
            period_str = (
                first_date if first_date == last_date else f"{first_date} - {last_date}"
            )
        else:
            period_str = "Н/Д"

        formatted_total = f"{total_spent:,.2f} ₽".replace(",", " ").replace(".", ",")

        history_items = []
        if "dt" in pat_df.columns:
            pat_df = pat_df.sort_values("dt", ascending=False)

        for _, row in pat_df.iterrows():
            date_label = (
                f"{MONTHS_RU.get(row['dt'].month, '')} { row['dt'].year}"
                if pd.notnull(row["dt"])
                else "Н/Д"
            )
            mes = row.get("Код Услуги", "Н/Д")
            formatted_summ = f"{row.get('Сумма', 0):,.2f} ₽".replace(",", " ").replace(
                ".", ","
            )

            item = html.Div(
                [
                    html.Div(
                        [
                            html.Div(
                                html.I(
                                    className="fas fa-check-circle",
                                    style={"color": "var(--primary)"},
                                ),
                                style={"marginRight": "15px", "fontSize": "20px"},
                            ),
                            html.Div(
                                [
                                    html.H6(
                                        f"Код МЭС: {mes}",
                                        style={
                                            "fontWeight": "700",
                                            "color": "var(--text-main)",
                                            "margin": "0 0 5px 0",
                                        },
                                    ),
                                    html.Div(
                                        [
                                            html.Span(
                                                [
                                                    html.I(
                                                        className="fas fa-calendar-day",
                                                        style={"marginRight": "6px"},
                                                    ),
                                                    date_label,
                                                ],
                                                style={
                                                    "marginRight": "15px",
                                                    "fontSize": "12px",
                                                    "fontWeight": "600",
                                                },
                                            ),
                                            html.Span(
                                                [
                                                    html.I(
                                                        className="fas fa-hospital",
                                                        style={"marginRight": "6px"},
                                                    ),
                                                    row.get(
                                                        "Наименование отделения", "Н/Д"
                                                    ),
                                                ],
                                                style={
                                                    "marginRight": "15px",
                                                    "fontSize": "12px",
                                                    "fontWeight": "600",
                                                },
                                            ),
                                        ],
                                        style={
                                            "display": "flex",
                                            "color": "var(--text-muted)",
                                        },
                                    ),
                                ],
                                style={"flex": 1},
                            ),
                            html.Div(
                                formatted_summ,
                                style={
                                    "fontWeight": "700",
                                    "color": "var(--text-main)",
                                },
                            ),
                        ],
                        style={"display": "flex", "alignItems": "center"},
                    )
                ],
                style={
                    "padding": "15px",
                    "borderBottom": "1px solid var(--grid-color)",
                },
            )
            history_items.append(item)

        modal_body = html.Div(
            [
                dbc.Row(
                    [
                        dbc.Col(
                            create_kpi_card(
                                "ОБЩИЙ ЧЕК",
                                formatted_total,
                                "fas fa-money-bill",
                                "#01B574",
                                "rgba(1,181,116,0.1)",
                            ),
                            width=3,
                        ),
                        dbc.Col(
                            create_kpi_card(
                                "УСЛУГ ОКАЗАНО",
                                f"{total_visits}",
                                "fas fa-stethoscope",
                                "#4318FF",
                                "rgba(67,24,255,0.1)",
                            ),
                            width=3,
                        ),
                        dbc.Col(
                            create_kpi_card(
                                "ОТДЕЛЕНИЙ",
                                f"{depts_visited}",
                                "fas fa-hospital",
                                "#FF7D00",
                                "rgba(255,125,0,0.1)",
                            ),
                            width=3,
                        ),
                        dbc.Col(
                            create_kpi_card(
                                "ПЕРИОД",
                                period_str,
                                "fas fa-calendar-alt",
                                "#39B8FF",
                                "rgba(57,184,255,0.1)",
                            ),
                            width=3,
                        ),
                    ],
                    style={"marginBottom": "30px"},
                ),
                html.Div(
                    [
                        html.H4(
                            [
                                html.I(
                                    className="fas fa-list-ul",
                                    style={
                                        "marginRight": "12px",
                                        "color": "var(--primary)",
                                    },
                                ),
                                "Детализация истории",
                            ],
                            style={"fontWeight": "800", "marginBottom": "20px"},
                        ),
                        html.Div(
                            history_items,
                            style={
                                "maxHeight": "400px",
                                "overflowY": "auto",
                                "border": "1px solid var(--grid-color)",
                                "borderRadius": "12px",
                            },
                        ),
                        html.Div(
                            [
                                html.Span(
                                    "ИТОГО ЗАРАБОТАНО ПО ИБ:",
                                    style={
                                        "fontWeight": "800",
                                        "fontSize": "14px",
                                        "color": "var(--text-muted)",
                                        "letterSpacing": "1px",
                                    },
                                ),
                                html.Span(
                                    formatted_total,
                                    style={
                                        "fontWeight": "900",
                                        "fontSize": "22px",
                                        "color": "#01B574",
                                        "marginLeft": "auto",
                                    },
                                ),
                            ],
                            style={
                                "display": "flex",
                                "alignItems": "center",
                                "padding": "20px 25px",
                                "marginTop": "20px",
                                "backgroundColor": "rgba(1, 181, 116, 0.05)",
                                "borderRadius": "16px",
                                "border": "2px solid #01B574",
                            },
                        ),
                    ],
                    style={
                        "backgroundColor": "var(--card-bg)",
                        "padding": "30px",
                        "borderRadius": "24px",
                        "boxShadow": "var(--shadow)",
                    },
                ),
            ]
        )

        return True, f"Карточка истории болезни: {patient_id}", modal_body
    return is_open, dash.no_update, dash.no_update


# --- SQL ПЕСОЧНИЦА ---
@app.callback(
    [
        Output("main-line-chart", "figure", allow_duplicate=True),
        Output("main-chart-title", "children", allow_duplicate=True),
        Output("sql-error", "children"),
        Output("smart-insights-container", "children", allow_duplicate=True),
    ],
    Input("btn-execute-sql", "n_clicks"),
    [State("sql-input", "value"), State("theme-store", "data")],
    prevent_initial_call=True,
)
def execute_custom_sql(n_clicks, query, theme):
    sql_insight = html.Div(
        [
            html.I(
                className="fas fa-bolt",
                style={"color": "#FF7D00", "marginRight": "12px", "fontSize": "20px"},
            ),
            html.Span(
                "DuckDB: ", style={"fontWeight": "800", "color": "var(--text-main)"}
            ),
            html.Span(
                "Запрос выполнен. Модальная детализация в этом режиме отключена.",
                style={"color": "var(--text-muted)"},
            ),
        ]
    )

    if not query or not query.strip():
        return dash.no_update, dash.no_update, "", dash.no_update

    try:
        df = get_optimized_data()
        safe_query = query.replace("[", '"').replace("]", '"')

        duckdb.register("medical_data", df)
        df_sql = duckdb.query(safe_query).df()

        if df_sql.empty:
            return (
                dash.no_update,
                dash.no_update,
                "✅ Запрос выполнен успешно (0 строк).",
                dash.no_update,
            )

        fig = go.Figure()
        x_col = df_sql.columns[0]
        max_y_value = 0
        color_idx = 0
        has_numeric = False

        colors = [
            {"hex": "#4318FF", "rgba": "rgba(67, 24, 255, 0.15)"},
            {"hex": "#FF7D00", "rgba": "rgba(255, 125, 0, 0.15)"},
            {"hex": "#01B574", "rgba": "rgba(1, 181, 116, 0.15)"},
        ]

        for col in df_sql.columns[1:]:
            if pd.api.types.is_numeric_dtype(df_sql[col]):
                has_numeric = True
                current_max = df_sql[col].max()
                if current_max > max_y_value:
                    max_y_value = current_max

                c = colors[color_idx % len(colors)]
                col_name_lower = str(col).lower()
                y_format = (
                    "%{y:,.2f} ₽"
                    if "сумм" in col_name_lower or "sum" in col_name_lower
                    else "%{y:,.0f}"
                )
                custom_data_sql = ["SQL|None"] * len(df_sql)

                fig.add_trace(
                    go.Scatter(
                        x=df_sql[x_col],
                        y=df_sql[col],
                        name=str(col),
                        mode="lines+markers",
                        line=dict(width=4, shape="linear", color=c["hex"]),
                        marker=dict(size=12, color=c["hex"]),
                        fill="tozeroy",
                        fillcolor=c["rgba"],
                        hovertemplate=f"<b>Ось X:</b> %{{x}}<br><b>{str(col)}:</b> {y_format}<extra></extra>",
                        customdata=custom_data_sql,
                    )
                )
                color_idx += 1

        if not has_numeric:
            return (
                dash.no_update,
                dash.no_update,
                "⚠️ Нужна хотя бы одна числовая колонка.",
                dash.no_update,
            )

        def format_label(label, max_len=35):
            s = str(label)
            if len(s) > max_len:
                words = s.split()
                if len(words) > 2:
                    s = (
                        " ".join(words[: len(words) // 2])
                        + "<br>"
                        + " ".join(words[len(words) // 2 :])
                    )
            if len(s) > max_len * 2:
                return s[: max_len * 2] + "..."
            return s

        x_range = [-0.5, len(df_sql) - 0.5]
        y_padding = max_y_value * 0.05 if max_y_value > 0 else 10
        tickvals = df_sql[x_col].tolist()
        ticktext = [format_label(v) for v in df_sql[x_col]]

        fig = apply_beautiful_layout(
            fig, theme, x_range=x_range, tickvals=tickvals, ticktext=ticktext
        )
        fig.update_yaxes(range=[-y_padding, max_y_value + y_padding])
        fig.update_layout(xaxis_tickangle=-30)

        return fig, "Аналитика (DuckDB Режим 🚀)", "", sql_insight
    except Exception as e:
        return (
            dash.no_update,
            dash.no_update,
            f"❌ Ошибка DuckDB: {str(e)}",
            dash.no_update,
        )


# --- ЭКСПОРТ EXCEL И КЛИЕНТСКИЕ КОЛЛБЭКИ ---
@app.callback(
    Output("download-xlsx", "data"),
    Input("btn-dl", "n_clicks"),
    [
        State("f-year", "value"),
        State("f-quarter", "value"),
        State("f-month", "value"),
        State("f-dept", "value"),
        State("f-profile", "value"),
        State("f-mes", "value"),
    ],
    prevent_initial_call=True,
)
def export_excel(n_clicks, years, quarters, months, depts, profiles, mes_list):
    df = get_optimized_data()
    mask = pd.Series(True, index=df.index)
    if years and "Year" in df.columns:
        mask &= df["Year"].isin(years)
    if quarters and "Quarter_Name" in df.columns:
        mask &= df["Quarter_Name"].isin(quarters)
    if months and "Month_Name" in df.columns:
        mask &= df["Month_Name"].isin(months)
    if depts and "Наименование отделения" in df.columns:
        mask &= df["Наименование отделения"].isin(depts)
    if profiles and "Наименование профиля" in df.columns:
        mask &= df["Наименование профиля"].isin(profiles)
    if mes_list and "Код Услуги" in df.columns:
        mask &= df["Код Услуги"].isin(mes_list)

    filtered_df = df[mask].copy()
    cols_to_drop = [
        "dt",
        "Year",
        "Month_Num",
        "Month_Name",
        "Quarter_Num",
        "Quarter_Name",
        "YearMonth",
        "Month_Str",
    ]
    columns_to_actually_drop = [
        col for col in cols_to_drop if col in filtered_df.columns
    ]
    filtered_df = filtered_df.drop(columns=columns_to_actually_drop)
    return dcc.send_data_frame(
        filtered_df.to_excel, "Clinical_Report_Export.xlsx", index=False
    )


app.clientside_callback(
    "function(n_clicks) { if (n_clicks) { return true; } return window.dash_clientside.no_update; }",
    Output("interactive-grid", "exportDataAsCsv"),
    Input("btn-export-grid", "n_clicks"),
    prevent_initial_call=True,
)

app.clientside_callback(
    """
    function(clickData) {
        if (!clickData) return window.dash_clientside.no_update;
        if (window.isShiftPressed) {
            return window.dash_clientside.no_update;
        }
        return clickData;
    }
    """,
    Output("filtered-click-data", "data"),
    Input("main-line-chart", "clickData"),
    prevent_initial_call=True
)

# 2. Логика сброса выделения
app.clientside_callback(
    """
    function(clickData, closeSmall, closeLarge) {
        const triggered = dash_clientside.callback_context.triggered.map(t => t.prop_id);

        function clearPlotlySelection() {
            const graphDiv = document.getElementById('main-line-chart');
            if (graphDiv && window.Plotly) {
                window.Plotly.restyle(graphDiv, 'selectedpoints', null);
            }
        }

        if (triggered.includes("btn-close-modal.n_clicks") || 
            triggered.includes("close-modal.n_clicks")) {
            clearPlotlySelection();
            return [null, null];
        }

        if (triggered.includes("main-line-chart.clickData")) {
            if (!window.isShiftPressed) {
                clearPlotlySelection();
                return [window.dash_clientside.no_update, null];
            }
        }
        
        return [window.dash_clientside.no_update, window.dash_clientside.no_update];
    }
    """,
    [
        Output("main-line-chart", "clickData"),
        Output("main-line-chart", "selectedData")
    ],
    [
        Input("main-line-chart", "clickData"),
        Input("btn-close-modal", "n_clicks"),
        Input("close-modal", "n_clicks")
    ],
    prevent_initial_call=True
)

app.clientside_callback(
    """
    function(n_clicks, current_style) {
        if (!n_clicks) return window.dash_clientside.no_update;
        
        let new_style = {...current_style};

        if (new_style.height === '500px') {
            new_style.height = '1300px'; 
        } else {
            new_style.height = '500px';
        }
        
        return new_style;
    }
    """,
    Output("sunburst-chart", "style"),
    Input("btn-expand-treemap", "n_clicks"),
    State("sunburst-chart", "style"),
    prevent_initial_call=True
)

app.clientside_callback(
    """
    function(selectedData) {
        console.log("[JS-3] Входящий selectedData:", selectedData, "| Shift:", window.isShiftPressed);
        if (!selectedData || !selectedData.points || selectedData.points.length === 0) {
            return null; 
        }
        
        if (window.isShiftPressed || selectedData.points.length > 1) {
            return selectedData;
        }
        
        return window.dash_clientside.no_update;
    }
    """,
    Output("filtered-selected-data", "data"),
    Input("main-line-chart", "selectedData"),
    prevent_initial_call=True
)

@app.callback(
    Output("heatmap-chart", "clickData"),
    Input("btn-reset-crossfilter", "n_clicks"),
    prevent_initial_call=True,
)
def reset_heatmap_selection(n_clicks):
    return None


app.clientside_callback(
    "function(n, c) { if(!n) return window.dash_clientside.no_update; const t = c === 'light' ? 'dark' : 'light'; document.documentElement.setAttribute('data-theme', t); return t; }",
    Output("theme-store", "data"),
    Input("theme-toggle", "n_clicks"),
    State("theme-store", "data"),
)


@app.callback(Output("theme-icon", "className"), Input("theme-store", "data"))
def update_icon(theme):
    return "fas fa-sun" if theme == "dark" else "fas fa-moon"


app.clientside_callback(
    "function(n) { if (n) { var e = document.getElementById('pdf-export-container'); var w = e.offsetWidth; var h = e.offsetHeight; html2pdf().set({margin: 20, filename: 'Clinical_Chart.pdf', image: { type: 'jpeg', quality: 1.0 }, html2canvas: { scale: 2, useCORS: true, logging: false }, jsPDF: { unit: 'px', format: [w + 40, h + 40], orientation: 'landscape' }}).from(e).save(); } return window.dash_clientside.no_update; }",
    Output("btn-pdf", "id"),
    Input("btn-pdf", "n_clicks"),
    prevent_initial_call=True,
)
app.clientside_callback(
    r"""function(sql_text) { if (!sql_text || sql_text.trim() === "") return ["", {"display": "none"}]; var warnings = []; var text = sql_text.trim(); if (!text.endsWith(';')) warnings.push("⚠️ Забыта точка с запятой ( ; ) в конце"); var openParens = (text.match(/\(/g) || []).length; var closeParens = (text.match(/\)/g) || []).length; if (openParens !== closeParens) warnings.push("⚠️ Незакрытые круглые скобки ( )"); var openBrackets = (text.match(/\[/g) || []).length; var closeBrackets = (text.match(/\]/g) || []).length; if (openBrackets !== closeBrackets) warnings.push("⚠️ Незакрытые квадратные скобки [ ]"); var singleQuotes = (text.match(/'/g) || []).length; if (singleQuotes % 2 !== 0) warnings.push("⚠️ Пропущена одинарная кавычка '"); var baseStyle = {"marginTop": "8px", "fontSize": "13px", "fontWeight": "600", "minHeight": "20px", "transition": "color 0.3s ease"}; if (warnings.length > 0) { baseStyle["color"] = "#FF7D00"; return [warnings.join("   |   "), baseStyle]; } else { baseStyle["color"] = "#01B574"; return ["✅ Синтаксис выглядит отлично!", baseStyle]; } }""",
    [Output("sql-modal-linter", "children"), Output("sql-modal-linter", "style")],
    Input("sql-modal-input", "value"),
    prevent_initial_call=False,
)


@app.callback(
    [
        Output("sql-editor-modal", "is_open"),
        Output("sql-modal-input", "value"),
        Output("sql-input", "value"),
    ],
    [
        Input("btn-expand-sql", "n_clicks"),
        Input("btn-save-sql-modal", "n_clicks"),
        Input("btn-close-sql-modal", "n_clicks"),
    ],
    [
        State("sql-editor-modal", "is_open"),
        State("sql-input", "value"),
        State("sql-modal-input", "value"),
    ],
    prevent_initial_call=True,
)
def toggle_sql_editor(exp, save, close, is_open, main_text, modal_text):
    ctx = dash.callback_context
    if not ctx.triggered:
        return is_open, dash.no_update, dash.no_update
    trig = ctx.triggered[0]["prop_id"].split(".")[0]
    if trig == "btn-expand-sql":
        return True, main_text, dash.no_update
    elif trig == "btn-save-sql-modal":
        return False, dash.no_update, modal_text
    elif trig == "btn-close-sql-modal":
        return False, dash.no_update, dash.no_update
    return is_open, dash.no_update, dash.no_update


@app.callback(
    [
        Output("presets-store", "data"),
        Output("dropdown-load-preset", "options"),
        Output("preset-msg", "children"),
    ],
    [Input("btn-save-preset", "n_clicks")],
    [
        State("input-preset-name", "value"),
        State("presets-store", "data"),
        State("f-year", "value"),
        State("f-quarter", "value"),
        State("f-month", "value"),
        State("f-dept", "value"),
        State("f-profile", "value"),
        State("f-mes", "value"),
    ],
    prevent_initial_call=True,
)
def save_preset(
    n_clicks, name, store_data, years, quarters, months, depts, profiles, mes
):
    if not name or name.strip() == "":
        return dash.no_update, dash.no_update, "❌ Введите название пресета!"
    store_data = store_data or {}
    store_data[name] = {
        "years": years,
        "quarters": quarters,
        "months": months,
        "depts": depts,
        "profiles": profiles,
        "mes": mes,
    }
    options = [{"label": k, "value": k} for k in store_data.keys()]
    return store_data, options, f"✅ Сценарий «{name}» успешно сохранен!"


@app.callback(
    [
        Output("f-year", "value"),
        Output("f-quarter", "value"),
        Output("f-month", "value"),
        Output("f-dept", "value"),
        Output("f-profile", "value"),
        Output("f-mes", "value"),
    ],
    [Input("dropdown-load-preset", "value")],
    [State("presets-store", "data")],
    prevent_initial_call=True,
)
def load_preset(selected_preset, store_data):
    if not selected_preset or not store_data or selected_preset not in store_data:
        return (
            dash.no_update,
            dash.no_update,
            dash.no_update,
            dash.no_update,
            dash.no_update,
            dash.no_update,
        )
    p = store_data[selected_preset]
    return (
        p.get("years"),
        p.get("quarters"),
        p.get("months"),
        p.get("depts"),
        p.get("profiles"),
        p.get("mes"),
    )

@app.callback(
    [
        Output("f-year", "value"), Output("f-quarter", "value"), Output("f-month", "value"),
        Output("f-dept", "value"), Output("f-profile", "value"), Output("f-mes", "value"),
        Output("f-patient", "value"), Output("f-yoy", "value")
    ],
    [Input("btn-reset-all-filters", "n_clicks")],
    prevent_initial_call=True
)
def clear_all_filter_inputs(n_clicks):
    if n_clicks:
        return None, None, None, None, None, None, "", []
    return dash.no_update


@app.callback(
    Output("selection-modal", "style"),
    Output("selection-modal-content", "children"),
    [
        Input("filtered-selected-data", "data"), 
        Input("btn-close-modal", "n_clicks")
    ],
    [State("f-group-by", "value")],
    prevent_initial_call=True
)
def show_selected_points_data(selected_data, close_clicks, group_by_col):
    ctx = dash.callback_context
    trigger = ctx.triggered[0]['prop_id'] if ctx.triggered else ""

    if "btn-close-modal" in trigger:
        return {"display": "none"}, ""

    if not selected_data or 'points' not in selected_data or len(selected_data['points']) == 0:
        return {"display": "none"}, ""

    df = get_optimized_data()
    if df.empty:
        return {"display": "none"}, ""

    conditions = []
    for p in selected_data['points']:
        dt_val = p.get('x')[:10] 
        c_data = str(p.get('customdata', '')) 
        
        if '|' in c_data:
            group_val = c_data.split('|')[1]
            safe_grp = str(group_val).replace("'", "''")
            conditions.append(f"(CAST(dt AS DATE) = '{dt_val}' AND \"{group_by_col}\" = '{safe_grp}')")

    if not conditions:
        return {"display": "none"}, ""

    where_points = " OR ".join(conditions)

    with duckdb.connect() as conn:
        conn.register('df', df)
        query = f'SELECT ROUND(SUM(Сумма), 2), COUNT(*) FROM df WHERE {where_points}'
        res = conn.execute(query).fetchone()

    total_sum = res[0] or 0
    total_count = res[1] or 0

    content = html.Div([
        html.Div([html.Span("Выделено точек: ", style={"color": "var(--text-muted)"}), html.B(f"{len(selected_data['points'])}")]),
        html.Div([html.Span("Количество услуг: ", style={"color": "var(--text-muted)"}), html.B(f"{total_count:,} ед.".replace(",", " "))]),
        html.Div([html.Span("Общая сумма: ", style={"color": "var(--text-muted)"}), html.B(f"{total_sum:,.2f} ₽".replace(",", " ").replace(".", ","), style={"color": "#01B574", "fontSize": "16px"})]),
    ])

    return {"display": "block"}, content

@app.callback(
    [
        Output("yoy-modal", "is_open"),
        Output("yoy-modal-title", "children"),
        Output("yoy-modal-body", "children"),
        Output("f-yoy", "value", allow_duplicate=True)
    ],
    [
        Input("f-yoy", "value"),
        Input("close-yoy-modal", "n_clicks")
    ],
    [
        State("yoy-modal", "is_open"),
        State("f-year", "value"),
        State("f-quarter", "value"),
        State("f-month", "value"),
        State("f-dept", "value"),
        State("f-profile", "value"),
        State("f-mes", "value"),
        State("f-patient", "value")
    ],
    prevent_initial_call=True
)
def toggle_yoy_comparison(yoy_val, close_clicks, is_open, years, quarters, months, depts, profiles, mes_list, patient):
    ctx = dash.callback_context
    trig = ctx.triggered[0]["prop_id"].split(".")[0]

    if trig == "close-yoy-modal":
        return False, dash.no_update, dash.no_update, [] 

    if trig == "f-yoy" and yoy_val and True in yoy_val:
        df = get_optimized_data()
        if df.empty:
            return True, "Сравнение с прошлым годом", html.Div("Нет данных в базе"), dash.no_update

        max_year_overall = int(df["Year"].max())
        target_year = max(years) if years else max_year_overall
        cy = target_year
        py = cy - 1

        where_base = build_where_clause(None, quarters, months, depts, profiles, mes_list, patient)

        with duckdb.connect() as conn:
            conn.register('df', df)

            query_totals = f'''
                SELECT 
                    "Year",
                    SUM(Сумма) as total_sum,
                    COUNT(DISTINCT "Номер ИБ") as total_patients,
                    COUNT(*) as total_mes
                FROM (
                    SELECT DISTINCT "Номер ИБ", "Код Услуги", "Year", Сумма 
                    FROM df 
                    WHERE {where_base} AND "Year" IN ({cy}, {py})
                )
                GROUP BY "Year"
            '''
            res_totals = conn.execute(query_totals).df()

            query_depts = f'''
                WITH dept_stats AS (
                    SELECT 
                        "Наименование отделения" as dept,
                        SUM(CASE WHEN "Year" = {cy} THEN Сумма ELSE 0 END) as sum_cy,
                        SUM(CASE WHEN "Year" = {py} THEN Сумма ELSE 0 END) as sum_py
                    FROM (
                        SELECT DISTINCT "Номер ИБ", "Код Услуги", "Наименование отделения", "Year", Сумма 
                        FROM df WHERE {where_base} AND "Year" IN ({cy}, {py})
                    )
                    GROUP BY "Наименование отделения"
                )
                SELECT dept, sum_cy, sum_py 
                FROM dept_stats 
                ORDER BY sum_cy DESC 
                LIMIT 3
            '''
            res_depts = conn.execute(query_depts).df()

            query_services = f'''
                WITH mes_stats AS (
                    SELECT 
                        "Код Услуги" as mes,
                        SUM(CASE WHEN "Year" = {cy} THEN Сумма ELSE 0 END) as sum_cy,
                        SUM(CASE WHEN "Year" = {py} THEN Сумма ELSE 0 END) as sum_py
                    FROM (
                        SELECT DISTINCT "Номер ИБ", "Код Услуги", "Year", Сумма 
                        FROM df WHERE {where_base} AND "Year" IN ({cy}, {py})
                    )
                    GROUP BY "Код Услуги"
                )
                SELECT mes, sum_cy, sum_py, (sum_cy - sum_py) as growth
                FROM mes_stats
                ORDER BY growth DESC 
                LIMIT 3
            '''
            res_services = conn.execute(query_services).df()

        stats = {cy: {"sum": 0, "pat": 0, "mes": 0}, py: {"sum": 0, "pat": 0, "mes": 0}}
        for _, row in res_totals.iterrows():
            y = int(row["Year"])
            if y in stats:
                stats[y] = {"sum": row["total_sum"], "pat": row["total_patients"], "mes": row["total_mes"]}

        avg_check_cy = stats[cy]["sum"] / stats[cy]["pat"] if stats[cy]["pat"] > 0 else 0
        avg_check_py = stats[py]["sum"] / stats[py]["pat"] if stats[py]["pat"] > 0 else 0

        def make_yoy_row(title, val_cy, val_py, is_currency=False):
            fmt = lambda x: f"{x:,.2f} ₽".replace(",", " ").replace(".", ",") if is_currency else f"{int(x):,.0f} шт.".replace(",", " ")
            diff = val_cy - val_py
            pct = (diff / val_py * 100) if val_py > 0 else 0
            
            color = "#01B574" if diff >= 0 else "#E11D48"
            bg_color = "rgba(1,181,116,0.1)" if diff >= 0 else "rgba(225,29,72,0.1)"
            icon = "fa-arrow-up" if diff >= 0 else "fa-arrow-down"
            sign = "+" if diff > 0 else ""
            
            return html.Div([
                html.Div([html.I(className="fas fa-chart-line", style={"marginRight": "10px", "color": "var(--primary)"}), title], style={"fontWeight": "800", "fontSize": "15px", "color": "var(--text-main)", "marginBottom": "15px"}),
                html.Div([
                    html.Div([html.Span(f"{py} год", style={"color": "var(--text-muted)", "fontSize": "12px", "display": "block", "fontWeight": "600"}), html.B(fmt(val_py), style={"fontSize": "20px", "color": "var(--text-main)"})], style={"flex": 1}),
                    html.Div([html.Span(f"{cy} год", style={"color": "var(--text-muted)", "fontSize": "12px", "display": "block", "fontWeight": "600"}), html.B(fmt(val_cy), style={"fontSize": "20px", "color": "var(--primary)"})], style={"flex": 1}),
                    html.Div([
                        html.Span([html.I(className=f"fas {icon}", style={"marginRight": "6px"}), f"{sign}{fmt(diff)}"], style={"display": "block", "fontWeight": "bold", "fontSize": "14px"}),
                        html.Span(f"({sign}{pct:.1f}%)", style={"fontSize": "12px", "fontWeight": "600"})
                    ], style={"flex": 1, "textAlign": "right", "color": color, "backgroundColor": bg_color, "padding": "10px 15px", "borderRadius": "12px"})
                ], style={"display": "flex", "alignItems": "center", "justifyContent": "space-between"})
            ], style={"backgroundColor": "var(--card-bg)", "padding": "25px", "borderRadius": "16px", "marginBottom": "15px", "border": "2px solid var(--grid-color)", "boxShadow": "var(--shadow)"})

        def make_compact_yoy_row(title, val_cy, val_py, is_currency=True):
            fmt = lambda x: f"{x:,.2f} ₽".replace(",", " ").replace(".", ",") if is_currency else f"{int(x):,.0f} шт.".replace(",", " ")
            diff = val_cy - val_py
            pct = (diff / val_py * 100) if val_py > 0 else 0
            
            color = "#01B574" if diff >= 0 else "#E11D48"
            bg_color = "rgba(1,181,116,0.1)" if diff >= 0 else "rgba(225,29,72,0.1)"
            icon = "fa-arrow-up" if diff >= 0 else "fa-arrow-down"
            sign = "+" if diff > 0 else ""

            return html.Div([
                html.Div(title, style={"fontWeight": "700", "fontSize": "13px", "color": "var(--text-main)", "marginBottom": "8px", "whiteSpace": "nowrap", "overflow": "hidden", "textOverflow": "ellipsis"}),
                html.Div([
                    html.Div([
                        html.Span(f"{py} год", style={"color": "var(--text-muted)", "fontSize": "11px", "display": "block"}), 
                        html.B(fmt(val_py), style={"fontSize": "14px", "color": "var(--text-main)"})
                    ], style={"flex": 1}),
                    html.Div([
                        html.Span(f"{cy} год", style={"color": "var(--text-muted)", "fontSize": "11px", "display": "block"}), 
                        html.B(fmt(val_cy), style={"fontSize": "14px", "color": "var(--primary)"})
                    ], style={"flex": 1}),
                    html.Div([
                        html.Span([html.I(className=f"fas {icon}", style={"marginRight": "4px"}), f"{sign}{fmt(diff)}"], style={"display": "block", "fontWeight": "bold", "fontSize": "12px"}),
                        html.Span(f"({sign}{pct:.1f}%)", style={"fontSize": "11px", "fontWeight": "600"})
                    ], style={"flex": 1, "textAlign": "right", "color": color, "backgroundColor": bg_color, "padding": "6px 10px", "borderRadius": "8px"})
                ], style={"display": "flex", "alignItems": "center", "justifyContent": "space-between"})
            ], style={"backgroundColor": "var(--card-bg)", "padding": "15px", "borderRadius": "12px", "marginBottom": "10px", "border": "1px solid var(--grid-color)"})

        dept_components = [
            html.H5("🏥 Топ-3 отделения (по выручке)", style={"fontWeight": "800", "color": "var(--text-main)", "marginTop": "25px", "marginBottom": "15px", "fontSize": "16px"})
        ]
        if res_depts.empty:
            dept_components.append(html.Div("Нет данных", style={"color": "var(--text-muted)", "fontSize": "13px"}))
        else:
            for _, row in res_depts.iterrows():
                dept_components.append(make_compact_yoy_row(str(row["dept"]), row["sum_cy"], row["sum_py"], is_currency=True))

        service_components = [
            html.H5("📈 Топ-3 быстрорастущие услуги (по приросту выручки)", style={"fontWeight": "800", "color": "var(--text-main)", "marginTop": "25px", "marginBottom": "15px", "fontSize": "16px"})
        ]
        if res_services.empty:
            service_components.append(html.Div("Нет данных", style={"color": "var(--text-muted)", "fontSize": "13px"}))
        else:
            for _, row in res_services.iterrows():
                mes_code = str(row['mes'])
                if mes_code.endswith('.0'):
                    mes_code = mes_code[:-2]
                
                service_components.append(make_compact_yoy_row(f"МЭС: {mes_code}", row["sum_cy"], row["sum_py"], is_currency=True))


        modal_body = html.Div([
            html.Div(f"Анализ проводится с учетом текущих фильтров. Год для сравнения выбран автоматически.", style={"color": "var(--text-muted)", "marginBottom": "20px", "fontSize": "13px"}),
            make_yoy_row("ВЫРУЧКА (ОБЩАЯ СУММА)", stats[cy]["sum"], stats[py]["sum"], is_currency=True),
            make_yoy_row("СРЕДНИЙ ЧЕК ПАЦИЕНТА", avg_check_cy, avg_check_py, is_currency=True),
            make_yoy_row("УНИКАЛЬНЫХ ПАЦИЕНТОВ", stats[cy]["pat"], stats[py]["pat"]),
            make_yoy_row("ОКАЗАННЫХ УСЛУГ (МЭС)", stats[cy]["mes"], stats[py]["mes"]),
            
            html.Div(dept_components),
            html.Div(service_components)
        ])

        return True, f"Сравнение {cy} vs {py} (YoY)", modal_body, dash.no_update

    return is_open, dash.no_update, dash.no_update, dash.no_update

def open_browser():
    webbrowser.open_new("http://127.0.0.1:8050")


if __name__ == "__main__":
    Timer(1.5, open_browser).start()
    app.run(debug=False, port=8050)
